from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, Response
from collections import OrderedDict
from functools import wraps
from dotenv import load_dotenv
import os, sqlite3, uuid as _uuid, json, csv, io, re
import requests as _requests
from scipy import stats as _scipy_stats
import numpy as _np
from datetime import datetime, timedelta
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from flask_wtf.csrf import CSRFProtect
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address

load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL", "").rstrip("/")
SUPABASE_KEY = os.getenv("SUPABASE_KEY", "")
_USE_SUPABASE = bool(SUPABASE_URL and SUPABASE_KEY)

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "tsl-dev-secret-2025")
app.config['WTF_CSRF_TIME_LIMIT'] = 3600
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50 MB upload limit
csrf = CSRFProtect(app)
limiter = Limiter(get_remote_address, app=app, default_limits=[], storage_uri="memory://")

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'portal.db')
PORTAL_FILES_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'portal_files')

GALLERY_FOLDER    = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'images', 'gallery')
GALLERY_DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'gallery_data.json')
APPS_EXTRA_FILE   = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'apps_extra.json')
PUBS_EXTRA_FILE   = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'pubs_extra.json')
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'avif', 'webp'}

def allowed_file(f): return '.' in f and f.rsplit('.',1)[1].lower() in ALLOWED_EXTENSIONS

def load_events():
    if not os.path.exists(GALLERY_DATA_FILE): return []
    with open(GALLERY_DATA_FILE,'r',encoding='utf-8') as f: return json.load(f)

def save_events(events):
    with open(GALLERY_DATA_FILE,'w',encoding='utf-8') as f: json.dump(events,f,ensure_ascii=False,indent=2)

def event_photos(key):
    folder = os.path.join(GALLERY_FOLDER, key)
    if not os.path.exists(folder): return []
    photos = [f for f in os.listdir(folder) if os.path.isfile(os.path.join(folder,f)) and allowed_file(f)]
    photos.sort(key=lambda f: os.path.getmtime(os.path.join(folder,f)))
    return photos

def load_extra_apps():
    if not os.path.exists(APPS_EXTRA_FILE): return []
    with open(APPS_EXTRA_FILE,'r',encoding='utf-8') as f: return json.load(f)

def save_extra_apps(data):
    with open(APPS_EXTRA_FILE,'w',encoding='utf-8') as f: json.dump(data,f,ensure_ascii=False,indent=2)

def load_extra_pubs():
    if not os.path.exists(PUBS_EXTRA_FILE): return []
    with open(PUBS_EXTRA_FILE,'r',encoding='utf-8') as f: return json.load(f)

def save_extra_pubs(data):
    with open(PUBS_EXTRA_FILE,'w',encoding='utf-8') as f: json.dump(data,f,ensure_ascii=False,indent=2)

def init_db():
    os.makedirs(PORTAL_FILES_DIR, exist_ok=True)
    if _USE_SUPABASE:
        return  # Tables managed in Supabase; skip local SQLite init
    with sqlite3.connect(DB_PATH) as conn:
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS projects (
                id TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                description TEXT,
                app_type TEXT,
                researcher_email TEXT NOT NULL,
                created_at TEXT DEFAULT (strftime('%Y-%m-%dT%H:%M:%S','now','localtime'))
            );
            CREATE TABLE IF NOT EXISTS project_participants (
                id TEXT PRIMARY KEY,
                project_id TEXT NOT NULL,
                code TEXT NOT NULL,
                gender TEXT,
                age INTEGER,
                group_name TEXT,
                enrolled_at TEXT DEFAULT (strftime('%Y-%m-%dT%H:%M:%S','now','localtime'))
            );
            CREATE TABLE IF NOT EXISTS measurements (
                id TEXT PRIMARY KEY,
                project_id TEXT NOT NULL,
                participant_id TEXT NOT NULL,
                phase TEXT,
                notes TEXT,
                data TEXT DEFAULT '{}',
                measured_at TEXT DEFAULT (strftime('%Y-%m-%dT%H:%M:%S','now','localtime'))
            );
            CREATE TABLE IF NOT EXISTS project_variables (
                id TEXT PRIMARY KEY,
                project_id TEXT NOT NULL,
                name TEXT NOT NULL,
                label TEXT,
                var_type TEXT DEFAULT 'number',
                unit TEXT
            );
            CREATE TABLE IF NOT EXISTS portal_files (
                id TEXT PRIMARY KEY,
                project_id TEXT,
                filename TEXT NOT NULL,
                original_name TEXT NOT NULL,
                size INTEGER DEFAULT 0,
                researcher_email TEXT NOT NULL,
                uploaded_at TEXT DEFAULT (strftime('%Y-%m-%dT%H:%M:%S','now','localtime'))
            );
            CREATE TABLE IF NOT EXISTS accounts (
                email TEXT PRIMARY KEY,
                password TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS news (
                id TEXT PRIMARY KEY,
                title TEXT NOT NULL,
                content TEXT,
                researcher_email TEXT NOT NULL,
                published INTEGER DEFAULT 1,
                created_at TEXT DEFAULT (strftime('%Y-%m-%dT%H:%M:%S','now','localtime'))
            );
            CREATE TABLE IF NOT EXISTS research_topics_extra (
                key TEXT PRIMARY KEY,
                title TEXT,
                summary TEXT,
                detail TEXT
            );
            CREATE TABLE IF NOT EXISTS contact_messages (
                id TEXT PRIMARY KEY,
                name TEXT,
                email TEXT,
                subject TEXT,
                message TEXT NOT NULL,
                created_at TEXT DEFAULT (strftime('%Y-%m-%dT%H:%M:%S','now','localtime'))
            );
            CREATE TABLE IF NOT EXISTS sessions (
                id TEXT PRIMARY KEY,
                project_id TEXT,
                participant_id TEXT,
                phase TEXT,
                notes TEXT,
                data TEXT DEFAULT '{}',
                received_at TEXT DEFAULT (strftime('%Y-%m-%dT%H:%M:%S','now','localtime'))
            );
            CREATE TABLE IF NOT EXISTS audit_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                researcher TEXT,
                action TEXT NOT NULL,
                target_table TEXT,
                target_id TEXT,
                detail TEXT,
                before_val TEXT,
                after_val TEXT,
                created_at TEXT DEFAULT (strftime('%Y-%m-%dT%H:%M:%S','now','localtime'))
            );
            CREATE TABLE IF NOT EXISTS project_collaborators (
                id TEXT PRIMARY KEY,
                project_id TEXT NOT NULL,
                researcher_email TEXT NOT NULL,
                role TEXT DEFAULT 'editor',
                invited_at TEXT DEFAULT (strftime('%Y-%m-%dT%H:%M:%S','now','localtime'))
            );
            CREATE TABLE IF NOT EXISTS reset_tokens (
                token TEXT PRIMARY KEY,
                email TEXT NOT NULL,
                expires_at TEXT NOT NULL,
                used INTEGER DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS project_protocols (
                id TEXT PRIMARY KEY,
                project_id TEXT NOT NULL,
                phase_name TEXT NOT NULL,
                description TEXT,
                due_offset_days INTEGER DEFAULT 0,
                sort_order INTEGER DEFAULT 0
            );
        ''')
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")
        # Seed default accounts via environment variables only
        existing = conn.execute("SELECT COUNT(*) FROM accounts").fetchone()[0]
        if existing == 0:
            seed_email = os.getenv("SEED_ADMIN_EMAIL","")
            seed_pw    = os.getenv("SEED_ADMIN_PASSWORD","")
            if seed_email and seed_pw:
                conn.execute("INSERT INTO accounts (email,password) VALUES (?,?)",
                             (seed_email, generate_password_hash(seed_pw)))
                conn.commit()
    # Migrations: add columns if missing
    _migrations = [
        ("measurements",    "ALTER TABLE measurements ADD COLUMN excluded INTEGER DEFAULT 0"),
        ("audit_log",       "ALTER TABLE audit_log ADD COLUMN before_val TEXT"),
        ("audit_log",       "ALTER TABLE audit_log ADD COLUMN after_val TEXT"),
    ]
    with sqlite3.connect(DB_PATH) as conn:
        for _, stmt in _migrations:
            try: conn.execute(stmt); conn.commit()
            except Exception: pass
        # Indexes
        for idx_stmt in [
            "CREATE INDEX IF NOT EXISTS idx_m_project      ON measurements(project_id)",
            "CREATE INDEX IF NOT EXISTS idx_m_participant  ON measurements(participant_id)",
            "CREATE INDEX IF NOT EXISTS idx_pp_project     ON project_participants(project_id)",
            "CREATE INDEX IF NOT EXISTS idx_pv_project     ON project_variables(project_id)",
            "CREATE INDEX IF NOT EXISTS idx_m_measured_at  ON measurements(measured_at)",
            "CREATE INDEX IF NOT EXISTS idx_audit_created  ON audit_log(created_at DESC)",
            "CREATE INDEX IF NOT EXISTS idx_proj_researcher ON projects(researcher_email)",
            "CREATE INDEX IF NOT EXISTS idx_collab_project  ON project_collaborators(project_id)",
        ]:
            try: conn.execute(idx_stmt)
            except Exception: pass
        # Unique index for participant code per project
        try:
            conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_pp_code ON project_participants(project_id, code)")
        except Exception: pass
        conn.commit()

def audit(action, table=None, target_id=None, detail=None, before=None, after=None):
    """Write an entry to the audit_log table. Silently ignores errors."""
    try:
        researcher = session.get("researcher", "system")
        before_val = json.dumps(before, ensure_ascii=False) if before is not None else None
        after_val  = json.dumps(after,  ensure_ascii=False) if after  is not None else None
        sb("POST", "audit_log", data={
            "id": str(_uuid.uuid4()),
            "researcher": researcher,
            "action": action,
            "target_table": table,
            "target_id": target_id,
            "detail": detail,
            "before_val": before_val,
            "after_val": after_val,
        })
    except Exception:
        pass

_SAFE_IDENT = re.compile(r'^[a-zA-Z_][a-zA-Z0-9_]*$')
_ALLOWED_TABLES = {
    'projects','project_participants','measurements','project_variables',
    'portal_files','accounts','news','research_topics_extra','contact_messages',
    'sessions','audit_log','project_collaborators','reset_tokens','project_protocols',
}

def _safe_col(col):
    col = col.strip()
    return col if (col == '*' or _SAFE_IDENT.match(col)) else None

def _parse_sb_params(params):
    filters, order_by, select_fields = {}, None, "*"
    if not params:
        return filters, order_by, select_fields
    for part in params.lstrip("?").split("&"):
        if not part or "=" not in part:
            continue
        key, val = part.split("=", 1)
        if key == "select":
            cols = [_safe_col(c) for c in val.split(",")]
            if all(cols):
                select_fields = ",".join(cols)
        elif key == "order":
            pieces = val.split(".")
            col = _safe_col(pieces[0])
            if col:
                direction = "DESC" if len(pieces) > 1 and pieces[1].lower() == "desc" else "ASC"
                order_by = f"{col} {direction}"
        elif _SAFE_IDENT.match(key) and val.startswith("eq."):
            filters[key] = val[3:]
    return filters, order_by, select_fields

def _sb_headers(prefer=None):
    h = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
    }
    if prefer:
        h["Prefer"] = prefer
    return h

def sb(method, table, data=None, params="", upsert=False):
    if table not in _ALLOWED_TABLES:
        print(f"DB error: disallowed table '{table}'")
        return []

    # ── Supabase REST API mode ──────────────────────────
    if _USE_SUPABASE:
        url = f"{SUPABASE_URL}/rest/v1/{table}{params}"
        try:
            if method == "GET":
                r = _requests.get(url, headers=_sb_headers(), timeout=10)
                r.raise_for_status()
                return r.json() or []
            elif method == "POST":
                d = dict(data or {})
                d.setdefault("id", str(_uuid.uuid4()))
                prefer = ("resolution=merge-duplicates,return=representation"
                          if upsert else "return=representation")
                r = _requests.post(url, headers=_sb_headers(prefer), json=d, timeout=10)
                r.raise_for_status()
                result = r.json()
                return result if isinstance(result, list) else ([result] if result else [])
            elif method == "PATCH":
                r = _requests.patch(url, headers=_sb_headers("return=minimal"),
                                    json=dict(data or {}), timeout=10)
                r.raise_for_status()
                return []
            elif method == "DELETE":
                r = _requests.delete(url, headers=_sb_headers(), timeout=10)
                r.raise_for_status()
                return []
        except Exception as e:
            print(f"Supabase error [{method} {table}]: {e}")
            return []

    # ── SQLite fallback (local dev) ─────────────────────
    filters, order_by, select_fields = _parse_sb_params(params)
    conn = None
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        if method == "GET":
            sql = f"SELECT {select_fields} FROM {table}"
            vals = []
            if filters:
                sql += " WHERE " + " AND ".join(f"{k}=?" for k in filters)
                vals = list(filters.values())
            if order_by:
                sql += f" ORDER BY {order_by}"
            rows = conn.execute(sql, vals).fetchall()
            result = []
            for row in rows:
                r = dict(row)
                if "data" in r and isinstance(r["data"], str):
                    try: r["data"] = json.loads(r["data"])
                    except: pass
                result.append(r)
            return result
        elif method == "POST":
            if not data: return []
            d = dict(data)
            d.setdefault("id", str(_uuid.uuid4()))
            if "data" in d and isinstance(d["data"], dict):
                d["data"] = json.dumps(d["data"], ensure_ascii=False)
            cols = list(d.keys())
            if upsert:
                non_pk = [c for c in cols if c != "id"]
                conn.execute(
                    f"INSERT INTO {table} ({','.join(cols)}) VALUES ({','.join('?'*len(cols))}) "
                    f"ON CONFLICT(id) DO UPDATE SET {', '.join(f'{c}=excluded.{c}' for c in non_pk)}",
                    [d[c] for c in cols])
            else:
                conn.execute(f"INSERT INTO {table} ({','.join(cols)}) VALUES ({','.join('?'*len(cols))})",
                             [d[c] for c in cols])
            conn.commit()
            row = conn.execute(f"SELECT * FROM {table} WHERE id=?", [d["id"]]).fetchone()
            if row:
                r = dict(row)
                if "data" in r and isinstance(r["data"], str):
                    try: r["data"] = json.loads(r["data"])
                    except: pass
                return [r]
            return []
        elif method == "PATCH":
            if not data or not filters: return []
            d = dict(data)
            if "data" in d and isinstance(d["data"], dict):
                d["data"] = json.dumps(d["data"], ensure_ascii=False)
            sql = f"UPDATE {table} SET {', '.join(f'{k}=?' for k in d)} WHERE {' AND '.join(f'{k}=?' for k in filters)}"
            conn.execute(sql, list(d.values()) + list(filters.values()))
            conn.commit()
            return []
        elif method == "DELETE":
            if not filters: return []
            sql = f"DELETE FROM {table} WHERE {' AND '.join(f'{k}=?' for k in filters)}"
            conn.execute(sql, list(filters.values()))
            conn.commit()
            return []
    except Exception as e:
        print(f"DB error [{method} {table}]: {e}")
        return []
    finally:
        if conn: conn.close()

def _sb_count(table, params=""):
    """Return total row count for a table. Used for pagination."""
    if _USE_SUPABASE:
        url = f"{SUPABASE_URL}/rest/v1/{table}{params if params else '?select=id'}"
        if "select=" not in url:
            url = url + ("&" if "?" in url else "?") + "select=id"
        try:
            r = _requests.get(url, headers={**_sb_headers(), "Prefer": "count=exact"}, timeout=10)
            cr = r.headers.get("Content-Range", "0/0")
            return int(cr.split("/")[-1]) if "/" in cr else 0
        except Exception:
            return 0
    else:
        try:
            with sqlite3.connect(DB_PATH) as conn:
                return conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
        except Exception:
            return 0

init_db()

def _get_account(email):
    rows = sb("GET", "accounts", params=f"?email=eq.{email}&select=password")
    return rows[0]["password"] if rows else None

def _set_account(email, password, already_hashed=False):
    hashed = password if already_hashed else generate_password_hash(password, method='pbkdf2:sha256')
    sb("POST", "accounts", data={"email": email, "password": hashed}, upsert=True)

def _account_exists(email):
    rows = sb("GET", "accounts", params=f"?email=eq.{email}&select=email")
    return bool(rows)

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "researcher" not in session: return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

def _require_project_owner(project_id):
    """Returns (project_dict, None) if OK, (None, redirect_response) if unauthorized."""
    proj = sb("GET","projects",params=f"?id=eq.{project_id}")
    if not proj or not isinstance(proj,list):
        flash("프로젝트를 찾을 수 없어요.")
        return None, redirect(url_for("portal"))
    owner = proj[0].get("researcher_email","")
    email = session.get("researcher","")
    # Allow owner OR collaborator with any role
    if owner != email:
        row = sb("GET", "project_collaborators",
                 params=f"?project_id=eq.{project_id}&researcher_email=eq.{email}&select=role")
        if not row:
            flash("이 프로젝트에 접근 권한이 없어요.")
            return None, redirect(url_for("portal"))
    return proj[0], None

# ─── Site-wide ───────────────────────────────────────────────
PROFESSOR = {
    "name":      "Yun-Jung Choi",
    "title":     "Professor",
    "bio":       "Professor Yun-Jung Choi is a nurse educator and researcher at the Red Cross College of Nursing, Chung-Ang University, Seoul, Korea.",
    "email":     "yunjungchoi@cau.ac.kr",
    "office":    "+82-2-820-6907",
    "lab_phone": "+82-2-820-5986",
    "address":   "Red Cross College of Nursing, Chung-Ang University, Seoul, South Korea",
    "photo":     "images/professor-placeholder.jpg",
}

RESEARCH_TOPICS = [
    {"key": "brainwave",   "title": "Brain Wave",              "summary": "Focused on Alpha waves to stabilise the mind, relieve stress, and improve learning efficiency.",         "image": "images/brain_wave.avif",  "gradient": "linear-gradient(135deg,#004d40 0%,#26a69a 100%)"},
    {"key": "simulation",  "title": "Simulation Education",    "summary": "Developing simulations using standardised patients and interactive PFA serious games.",                   "image": "images/simulation.avif",  "gradient": "linear-gradient(135deg,#1a237e 0%,#5c6bc0 100%)"},
    {"key": "binaural",    "title": "Binaural Beat",           "summary": "Developing binaural beats and modulating brainwaves to promote mental wellbeing.",                       "image": "images/binaural.avif",    "gradient": "linear-gradient(135deg,#4a148c 0%,#ab47bc 100%)"},
    {"key": "app",         "title": "Developing Applications", "summary": "Building evidence-based mobile apps for disaster survivors and healthcare workers.",                     "image": "images/app.avif",         "gradient": "linear-gradient(135deg,#0F6B6B 0%,#1ec0c0 100%)"},
]

TEAM = {
    "phd": [
        {"name": "Hae-In Namgung",  "interest": "Simulation education, intervention research, RCT, PTSD, Artificial intelligence (AI)", "image": "images/member-haein-placeholder.jpg"},
        {"name": "Jae-Won Kwak",    "interest": "Community mental health nursing, community addiction management",                        "image": "images/member-jaewon-placeholder.jpg"},
        {"name": "Joo-Young Jin",   "interest": "Nursing simulation, disaster nursing, psychological safety",                            "image": "images/member-jooyoung-placeholder.jpg"},
    ],
    "ma": [
        {"name": "Chae-Young Lee",  "interest": "Anxiety, PTSD, depression, suicide, addiction",                                       "image": "images/member-chaeyoung-placeholder.jpg"},
    ],
    "alumni": [
        {"name": "Run-Ju Choi",   "image": "images/run-ju_choi.avif",    "current_position": ""},
        {"name": "Eun-Jung Ko",   "image": "images/eunjung_ko.avif",     "current_position": ""},
        {"name": "Younjoo Um",    "image": "images/younjoo_um.avif",     "current_position": ""},
        {"name": "Dong-Hee Cho",  "image": "images/dong-hee_cho.avif",   "current_position": ""},
        {"name": "Hee-Won Song",  "image": "images/hee-won_song.avif",   "current_position": ""},
    ],
}

APPS = [
    {
        "key":       "pls",
        "icon":      "🫀",
        "gradient":  "linear-gradient(135deg,#00695c 0%,#26a69a 100%)",
        "name":      "PLS",
        "full_name": "Psychological Life Support",
        "label":     "Mobile PFA Application",
        "desc_kr":   "재난구호자를 위한 심리적인명구조술 PLS는\n재난상황에 대한 정보제공 및 심리적응급처치 기술 적용,\n재난 후 구호요원들의 마음회복을 돕기위한 앱입니다.",
        "desc_en":   "PLS (Psychological Life Skill) for disaster relief workers is providing information on disaster situations and application of psychological first aid techniques. It is a mobile app to help rescue workers recover their minds after a disaster.",
        "tags":      ["Disaster Workers", "Mental Health", "Psychological First Aid"],
        "screenshot": None,
        "content_img_kr": "images/pls_kr_content.avif",
        "content_img_en": "images/pls_en_content.avif",
        "dl_kr":  "https://m.onestore.co.kr/ko-kr/apps/appsDetail.omp?prodId=0000747597",
        "dl_en":  "https://m.onestore.co.kr/ko-kr/apps/appsDetail.omp?prodId=0000747600",
        "subsections": [],
    },
    {
        "key":       "tls",
        "icon":      "🌿",
        "gradient":  "linear-gradient(135deg,#1b5e20 0%,#66bb6a 100%)",
        "name":      "TLS",
        "full_name": "Training of Life Skills",
        "label":     "Mobile PFA Application",
        "desc_kr":   "재난경험자를 위한 마음치유기술 TLS (Training for Life Skills)는\n재난피해로 심리적 상처를 입은 분들의 심리회복을 돕기 위한 앱입니다.",
        "desc_en":   "TLS (Training for Life Skills) for disaster experienced people is an app to help the psychological recovery of those who have been psychologically wounded by disaster damage.",
        "tags":      ["Disaster Survivors", "Trauma Recovery", "Skills Training"],
        "screenshot": None,
        "content_img_kr": "images/tls_kr_content.avif",
        "content_img_en": "images/tls_en_content.avif",
        "dl_kr":  "https://m.onestore.co.kr/ko-kr/apps/appsDetail.omp?prodId=0000747598",
        "dl_en":  "https://m.onestore.co.kr/ko-kr/apps/appsDetail.omp?prodId=0000747605",
        "subsections": [],
    },
    {
        "key":       "mind_therapy",
        "icon":      "🧠",
        "gradient":  "linear-gradient(135deg,#311b92 0%,#9c27b0 100%)",
        "name":      "Mind Therapy",
        "full_name": "Neurofeedback Stress Management",
        "label":     "",
        "desc_kr":   "",
        "desc_en":   "The 'Mind Therapy' application is designed to help people recover from traumatic events. Developed by applying EEG and heart rate variability technology, it can measure left-right brain balance, brain activity, and stress recovery through real-time EEG/heart rate variability tests.",
        "tags":      ["Neurofeedback", "Binaural Beat", "Stress & PTSD"],
        "screenshot": "images/app_screenshot3.png",
        "content_img_kr": None,
        "content_img_en": None,
        "dl_kr":  None,
        "dl_en":  None,
        "subsections": [
            {
                "title": "Brain Therapy",
                "image": "images/mt_brain_therapy.avif",
                "desc":  "We offer a meditation program with neurofeedback and binaural beats.",
                "links": [
                    {"text": "Neurofeedback",  "url": None},
                    {"text": "Binaural beats", "url": None},
                ],
            },
            {
                "title": "Autonomic Nerve Therapy",
                "image": "images/mt_autonomic.avif",
                "desc":  "The deep breathing and butterfly hug programs are prepared for your autonomic nervous system stabilization. Several options for deep breathing are provided — choose whatever you feel comfortable with. A video guide for the butterfly hug is also included.",
                "links": [],
            },
            {
                "title": "Brain Game",
                "image": "images/mt_brain_game.avif",
                "desc":  "We visually represent the response when alpha waves are enhanced to help train the brain waves. The balloon's face will change along with your alpha waves. Enhance alpha waves through an exciting game!",
                "links": [],
            },
        ],
    },
]

# ─── Publications ────────────────────────────────────────────
ARTICLES = [
    # 2024
    {"year": 2024, "title": "Feasibility of a Mobile App for Traumatic Stress Management Using Neurofeedback-based Mediation and Binaural Beat Music: A Pilot Randomized Controlled Trial", "authors": "Choi, Y. J., Cho, D. H. & Lee, N. R.", "journal": "Digital Health, 10", "doi": "https://doi.org/10.1177/20552076241308986"},
    {"year": 2024, "title": "Mental Health Status and Related Factors of Citizens 6 Months after Mass Death and Injury Due to Crowd Crush Incident: Focused on the Itaewon Disaster in 2022", "authors": "Choi, Y. J., Song, H., Namgung, H. I. & Lee, N. R.", "journal": "Disaster Medicine and Public Health Preparedness, 19, e11", "doi": "https://doi.org/10.1017/dmp.2024.342"},
    {"year": 2024, "title": "Mediating effect of bicultural acceptability among multicultural adolescents in the relationship between depression and life satisfaction", "authors": "Choi, Y. J. & Um, Y. J.", "journal": "Archives of Psychiatric Nursing, 53", "doi": "https://doi.org/10.1016/j.apnu.2024.10.003"},
    {"year": 2024, "title": "Still in there—citizens' well-being and PTSD after Seoul Halloween crowd crush in Korea: A cross-sectional study", "authors": "Choi, Y. J. & Namgung, H. I.", "journal": "Scientific Reports, 14, 20537", "doi": "https://doi.org/10.1038/s41598-024-71631-9"},
    {"year": 2024, "title": "Development and Effect of an Interactive Simulated Education Program for Psychological First Aid: A Randomized Controlled Trial", "authors": "Choi, E. J. & Choi, Y. J.", "journal": "Journal of Nursing Management, 8806047", "doi": "https://doi.org/10.1155/2024/8806047"},
    {"year": 2024, "title": "Experiences of Family Caregiver of Older People with Dementia in Korea during the COVID-19 Pandemic: A Qualitative Analysis", "authors": "Joh, E. S. & Choi, Y. J.", "journal": "Journal of Gerontological Nursing, 50(10)", "doi": "https://doi.org/10.3928/00989134-20240916-01"},
    {"year": 2024, "title": "The Association between Fear of COVID-19, Obsession with COVID-19, and Post Traumatic Stress Disorder in Korean Emergency Rescue Firefighters", "authors": "Choi, Y. J. & Song, H.", "journal": "International Journal of Mental Health Promotion, 26(6)", "doi": "https://doi.org/10.32604/ijmhp.2024.050824"},
    {"year": 2024, "title": "Citizens' Mental Health Issues and Psychological Trauma Experience due to a Crowd-Crush Disaster in Korea", "authors": "Choi, Y. J., Kwak, J. W. & Namgung, H. I.", "journal": "International Journal of Mental Health Promotion, 26(6)", "doi": "https://doi.org/10.32604/ijmhp.2024.050458"},
    {"year": 2024, "title": "Efficacy of a virtual nursing simulation-based education to provide psychological support for patients affected by infectious disease disasters: a randomized controlled trial", "authors": "Ko, E. & Choi, Y. J.", "journal": "BMC Nursing, 23, 230", "doi": "https://doi.org/10.1186/s12912-024-01901-4"},
    {"year": 2024, "title": "Effectiveness of a fire disaster PFA simulation game: A single-blinded trial", "authors": "Choi, Y. J. & Song, H.", "journal": "Disaster Medicine and Public Health Preparedness, 18, e64", "doi": "https://doi.org/10.1017/dmp.2024.47"},
    {"year": 2024, "title": "The Development, Implementation, and Evaluation of a Geriatric Disaster Nursing Simulation Intervention With Supportive Debriefing", "authors": "Jin, J. Y. & Choi, Y. J.", "journal": "Simulation in Healthcare, 19(5)", "doi": "https://doi.org/10.1097/SIH.0000000000000780"},
    # 2023
    {"year": 2023, "title": "What Influenced Frontline Nurses' Mental Health During the Early Phase of the Covid-19 Pandemic", "authors": "Choi, Y. J., Um, Y. J. & Cho, D. H.", "journal": "International Nursing Review, 70(4)", "doi": "https://doi.org/10.1111/inr.12895"},
    {"year": 2023, "title": "Development of a Multiple-Patient Simulation and its Effectiveness in Clinical Judgment and Practice Readiness: A Randomized Controlled Trial", "authors": "Namgung, H. I., Choi, Y. J. & Kang, J. S.", "journal": "Clinical Simulation in Nursing, 83, 101448", "doi": "https://doi.org/10.1016/j.ecns.2023.101448"},
    {"year": 2023, "title": "Effects of a web-based education for community mental health case managers on physical healthcare for clients with severe mental illness", "authors": "Lee, J. & Choi, Y. J.", "journal": "AIMS Public Health, 10(3)", "doi": "https://doi.org/10.3934/publichealth.2023045"},
    {"year": 2023, "title": "Neurofeedback Effect on Symptoms of Posttraumatic Stress Disorder: A Systematic Review and Meta-Analysis", "authors": "Choi, Y. J., Choi, E. J. & Ko, E.", "journal": "Applied Psychophysiology and Biofeedback, 48", "doi": "https://doi.org/10.1007/s10484-023-09593-3"},
    {"year": 2023, "title": "Inpatient meditation for alcohol use disorder reduces mood dysregulation: A pilot study", "authors": "Choi, Y. J., Cho, D. H. & Lee, N. R.", "journal": "Social Behavior and Personality, 51(10)", "doi": "https://doi.org/10.2224/sbp.12451"},
    {"year": 2023, "title": "Professional quality of life, resilience, posttraumatic stress and leisure activity among intensive care unit nurses", "authors": "Shin, N. & Choi, Y. J.", "journal": "International Nursing Review", "doi": "https://doi.org/10.1111/inr.12850"},
    {"year": 2023, "title": "Effects of a mental health nursing simulation for general ward nurses: A pilot study", "authors": "Lee, M. Y. & Choi, Y. J.", "journal": "Nursing Open, 10(5)", "doi": None},
    {"year": 2023, "title": "Topic Models to Analyze Disaster-Related Newspaper Articles: Focusing on COVID-19", "authors": "Choi, Y. J. & Um, Y. J.", "journal": "International Journal of Mental Health Promotion, 25(3)", "doi": None},
    {"year": 2023, "title": "Disaster Healthcare Workers' Experience of Using the Psychological First Aid Mobile App During Disaster Simulation Training", "authors": "Choi, Y. J., Jung, H. S., Choi, E. J. & Ko, E.", "journal": "Disaster Medicine and Public Health Preparedness, 17, e55", "doi": None},
    {"year": 2023, "title": "The early emotional responses and central issues of people in the epicenter of the COVID-19 pandemic: An analysis from twitter text mining", "authors": "Choi, E. J. & Choi, Y. J.", "journal": "International Journal of Mental Health Promotion", "doi": None},
    # 2022
    {"year": 2022, "title": "Student nurse experiences in public healthcare clinical practice during the COVID-19 pandemic: A qualitative study", "authors": "Choi, Y. J. & Um, Y. J.", "journal": "Nurse Education Today, 119, 105586", "doi": None},
    {"year": 2022, "title": "The effects of a home-visit nursing simulation for older people with dementia on nursing students' communication skills, self-efficacy, and critical thinking propensity", "authors": "Choi, Y. J. & Um, Y. J.", "journal": "Nurse Education Today, 119, 105564", "doi": None},
    {"year": 2022, "title": "Standardized patient experiences study on clinical performance evaluation of nursing college students' ability: A qualitative study", "authors": "Choi, Y. J., Won, M. R. & Yoo, S. Y.", "journal": "Nurse Education Today, 118, 105437", "doi": None},
    {"year": 2022, "title": "Effects of Nursing Care Using Binaural Beat Music on Anxiety, Pain, and Vital Signs in Surgery Patients", "authors": "Jang, Y. & Choi, Y. J.", "journal": "Journal of PeriAnesthesia Nursing, 37(6)", "doi": None},
    {"year": 2022, "title": "A grounded theory on school nursing experiences with major pandemic diseases", "authors": "Um, Y. J. & Choi, Y. J.", "journal": "INQUIRY: The Journal of Health Care Organization, 59", "doi": None},
    {"year": 2022, "title": "A simulation-based nursing education of psychological first aid for adolescents exposed to hazardous chemical disasters", "authors": "Kim, H. W. & Choi, Y. J.", "journal": "BMC Medical Education, 22(1)", "doi": None},
    {"year": 2022, "title": "Simulation-based education for nurses in caring for the psychological well-being of survivors of disaster", "authors": "Yun, S. M. & Choi, Y. J.", "journal": "The Journal of Continuing Education in Nursing, 53(3)", "doi": None},
    {"year": 2022, "title": "Clinical Nurses' Continuing Education Needs in Acute Burn Care", "authors": "Oh, D. & Choi, Y. J.", "journal": "The Journal of Continuing Education in Nursing, 53(2)", "doi": None},
    {"year": 2022, "title": "Efficacy of a Community-Based Trauma Recovery Program after a Fire Disaster", "authors": "Choi, Y. J., Won, M. R. & Cho, D. H.", "journal": "International Journal of Mental Health Promotion, 24", "doi": None},
    {"year": 2022, "title": "The effect of a simulated fire disaster psychological first aid training program on the self-efficacy, competence, and knowledge of mental health practitioners", "authors": "Park, J. S. & Choi, Y. J.", "journal": "Disaster Medicine and Public Health Preparedness, 16(1)", "doi": None},
    # 2021
    {"year": 2021, "title": "COVID-19 and risk factors of anxiety and depression in South Korea", "authors": "Hyun, J. et al.", "journal": "Psychiatry Investigation, 18(9)", "doi": None},
    {"year": 2021, "title": "Explicit and implicit attitudes toward people with COVID-19: Need for community mental health services", "authors": "Choi, Y. J. & Cho, D. H.", "journal": "Social Behavior and Personality, 49(11)", "doi": None},
    {"year": 2021, "title": "Managing traumatic stress using a mental health care mobile app: A pilot study", "authors": "Choi, Y. J., Ko, E. J., Choi, E. J. & Um, Y. J.", "journal": "International Journal of Mental Health Promotion, 23", "doi": None},
    {"year": 2021, "title": "The mediating effect of life satisfaction and the moderated mediating effect of social support on the relationship between depression and suicidal behavior among older adults", "authors": "Won, M. R., Choi, E. J., Ko, E., Um, Y. J. & Choi, Y. J.", "journal": "International Journal of Geriatric Psychiatry, 36(11)", "doi": None},
    {"year": 2021, "title": "Effects of Stress, Depression, and Problem Drinking on Suicidal Ideation among Korean Workers", "authors": "Choi, Y. J., Won, M. R. & Um, Y. J.", "journal": "International Journal of Mental Health Promotion, 23", "doi": None},
    {"year": 2021, "title": "Nursing students' extracurricular activity experiences of suicide prevention volunteering: A qualitative study", "authors": "Yoo, S. Y., Choi, E. J. & Choi, Y. J.", "journal": "Nurse Education Today, 102", "doi": None},
    {"year": 2021, "title": "Effects of a psychological first aid simulated training for pregnant flood victims on disaster relief worker's knowledge, competence, and self-efficacy", "authors": "Kang, J. Y. & Choi, Y. J.", "journal": "Applied Nursing Research, 57", "doi": None},
    # 2020
    {"year": 2020, "title": "The effect of employment status on people with a mental disability and on daily life satisfaction", "authors": "Um, Y. J. & Choi, Y. J.", "journal": "Current Psychology", "doi": None},
    {"year": 2020, "title": "Challenges and growth as a mental health professional from volunteering experiences in the community gambling awareness campaign", "authors": "Yoo, S. Y., Choi, Y. J. & Um, Y. J.", "journal": "International Journal of Mental Health Promotion, 22(2)", "doi": None},
    {"year": 2020, "title": "Debriefing model for psychological safety in nursing simulations: a qualitative study", "authors": "Ko, E. & Choi, Y. J.", "journal": "International Journal of Environmental Research and Public Health, 17(8)", "doi": None},
    {"year": 2020, "title": "The Mediating Role of Job Satisfaction in the Relationship between Disaster Relief Workers' Perception of Survivors' Rights and Their Performance of Human Rights Advocacy", "authors": "Choi, Y. J. & Ko, E.", "journal": "International Journal of Mental Health Promotion, 22", "doi": None},
    {"year": 2020, "title": "Bilingual gatekeepers' experiences of immigrant women's acculturative stress and mental health improvement in Korea: A qualitative analysis", "authors": "Choi, Y. J.", "journal": "Social Behavior and Personality, 48(9)", "doi": None},
    {"year": 2020, "title": "Psychological first-aid experiences of disaster health care workers: a qualitative analysis", "authors": "Choi, Y. J.", "journal": "Disaster Medicine and Public Health Preparedness, 14(4)", "doi": None},
    # 2019
    {"year": 2019, "title": "Relationships between smartphone dependency and aggression among middle school students", "authors": "Um, Y. J., Choi, Y. J. & Yoo, S. Y.", "journal": "International Journal of Environmental Research and Public Health, 16(19)", "doi": None},
    {"year": 2019, "title": "Nurses' Positive Experiences in Caring for Older Adults With Dementia: A Qualitative Analysis", "authors": "Choi, Y. J. & Choi, H. B.", "journal": "Journal of Gerontological Nursing, 45(1)", "doi": None},
    {"year": 2019, "title": "Nursing competency and educational needs for clinical practice of Korean nurses", "authors": "Kim, S. O. & Choi, Y. J.", "journal": "Nurse Education in Practice, 34", "doi": None},
    # 2018
    {"year": 2018, "title": "Relationships of substance use and sexual behavior of female junior high school students in Korea", "authors": "Lee, G. Y., Song, S. H. & Choi, Y. J.", "journal": "Journal of Child & Adolescent Substance Abuse, 27(5-6)", "doi": None},
    {"year": 2018, "title": "Associations among elder abuse, depression and PTSD in South Korean older adults", "authors": "Choi, Y. J., O'Donnell, M., Choi, H. B., Jung, H. S. & Cowlishaw, S.", "journal": "International Journal of Environmental Research and Public Health, 15(9)", "doi": None},
    {"year": 2018, "title": "Three-dimensional needs of standardized patients in nursing simulations and collaboration strategies: A qualitative analysis", "authors": "Jin, H. R. & Choi, Y. J.", "journal": "Nurse Education Today, 68", "doi": None},
    {"year": 2018, "title": "Disaster reintegration model: a qualitative analysis on developing Korean disaster mental health support model", "authors": "Choi, Y. J., Choi, H. B. & O'Donnell, M.", "journal": "International Journal of Environmental Research and Public Health, 15(2)", "doi": None},
    {"year": 2018, "title": "The value of psychosocial group activity in nursing education: A qualitative analysis", "authors": "Choi, Y. J.", "journal": "Nurse Education Today, 64", "doi": None},
    # 2017
    {"year": 2017, "title": "Effects of a program to improve mental health literacy for married immigrant women in Korea", "authors": "Choi, Y. J.", "journal": "Archives of Psychiatric Nursing, 31(4)", "doi": None},
    {"year": 2017, "title": "Analysis of Korean adolescents' sexual experience and substance use", "authors": "Lee, G. Y. & Choi, Y. J.", "journal": "Social Behavior and Personality, 45(5)", "doi": None},
    {"year": 2017, "title": "Family stress and coping from hospitalization of clients with severe alcohol use disorder in Korea", "authors": "Park, G. H. & Choi, Y. J.", "journal": "Journal of Addictions Nursing, 28(1)", "doi": None},
    {"year": 2017, "title": "Factors associated with perceived depression of Korean adults: secondary data from the Korean Community Health Survey", "authors": "Won, M. R., Ahn, M. S. & Choi, Y. J.", "journal": "Community Mental Health Journal, 53", "doi": None},
    {"year": 2017, "title": "Undergraduate nursing student mentors' experiences of peer mentoring in Korea: A qualitative analysis", "authors": "Won, M. R. & Choi, Y. J.", "journal": "Nurse Education Today, 51", "doi": None},
    # 2016
    {"year": 2016, "title": "Immigrant women's acculturation stress and coping strategies in Korea: A qualitative analysis", "authors": "Choi, Y. J.", "journal": "International Journal of Intercultural Relations, 55", "doi": None},
    {"year": 2016, "title": "Evaluation of a program on self-esteem and ego-identity for Korean nursing students", "authors": "Choi, Y. J.", "journal": "Nursing & Health Sciences, 18(3)", "doi": None},
    {"year": 2016, "title": "Associations among acculturation stress, mental health literacy, and mental health of married immigrant women in Korea", "authors": "Choi, Y. J. & Park, G. H.", "journal": "International Journal of Mental Health Promotion, 18(4)", "doi": None},
    {"year": 2016, "title": "Undergraduate students' experiences of an integrated psychiatric nursing curriculum in Korea", "authors": "Choi, Y. J.", "journal": "Issues in Mental Health Nursing, 37(8)", "doi": None},
    {"year": 2016, "title": "Effects of an obesity management mentoring program for Korean children", "authors": "Lee, G. Y. & Choi, Y. J.", "journal": "Applied Nursing Research, 31", "doi": None},
    {"year": 2016, "title": "Mental health problems and acculturative issues among married immigrant women in Korea: A qualitative study", "authors": "Choi, Y. J.", "journal": "Women & Health, 56(6)", "doi": None},
    # 2015
    {"year": 2015, "title": "Association of school, family, and mental health characteristics with suicidal ideation among Korean adolescents", "authors": "Lee, G. Y. & Choi, Y. J.", "journal": "Research in Nursing & Health, 38(4)", "doi": None},
    {"year": 2015, "title": "Mobile phone overuse among elementary school students in Korea: Factors associated with mobile phone use as a behavior addiction", "authors": "Kim, R., Lee, K. J. & Choi, Y. J.", "journal": "Journal of Addictions Nursing, 26(2)", "doi": None},
    {"year": 2015, "title": "The impact of gender, culture, and society on Korean women's mental health", "authors": "Choi, Y. J.", "journal": "Social Behavior and Personality, 43(4)", "doi": None},
    {"year": 2015, "title": "Efficacy of adjunctive treatments added to olanzapine or clozapine for weight control in patients with schizophrenia: a systematic review and meta-analysis", "authors": "Choi, Y. J.", "journal": "The Scientific World Journal, 2015", "doi": None},
    # 2014 and earlier
    {"year": 2014, "title": "Phases of shopping addiction evidenced by experiences of compulsive buyers", "authors": "Sohn, S. H. & Choi, Y. J.", "journal": "International Journal of Mental Health and Addiction, 12", "doi": None},
    {"year": 2013, "title": "Are Koreans prepared for the rapid increase of the single-household elderly? Life satisfaction and depression of the single-household elderly in Korea", "authors": "Won, M. R. & Choi, Y. J.", "journal": "The Scientific World Journal, 2013", "doi": None},
    {"year": 2013, "title": "Standardized patients for Korean psychiatric nursing student simulations", "authors": "Choi, Y. J.", "journal": "Clinical Simulation in Nursing, 9(9)", "doi": None},
    {"year": 2013, "title": "A pilot study on effects of a group program using recreational therapy to improve interpersonal relationships for undergraduate nursing students", "authors": "Choi, Y. J. & Won, M. R.", "journal": "Archives of Psychiatric Nursing, 27(1)", "doi": None},
    {"year": 2012, "title": "A model of compulsive buying: Dysfunctional beliefs and self-regulation of compulsive buyers", "authors": "Sohn, S. H. & Choi, Y. J.", "journal": "Social Behavior and Personality, 40(10)", "doi": None},
    {"year": 2012, "title": "Exploring experiences of psychiatric nursing simulations using standardized patients for undergraduate students", "authors": "Choi, Y. J.", "journal": "Asian Nursing Research, 6(3)", "doi": None},
    {"year": 2012, "title": "Effects of an emotion management nursing program for patients with schizophrenia", "authors": "Won, M. R., Lee, K. J., Lee, J. H. & Choi, Y. J.", "journal": "Archives of Psychiatric Nursing, 26(1)", "doi": None},
    {"year": 2010, "title": "The effect of an anger management program for family members of patients with alcohol use disorders", "authors": "Son, J. Y. & Choi, Y. J.", "journal": "Archives of Psychiatric Nursing, 24(1)", "doi": None},
    {"year": 2009, "title": "Efficacy of treatments for patients with obsessive-compulsive disorder: A systematic review", "authors": "Choi, Y. J.", "journal": "Journal of the American Academy of Nurse Practitioners, 21(4)", "doi": None},
    {"year": 2008, "title": "Experiences and challenges of informal caregiving for Korean immigrants", "authors": "Han, H. R., Choi, Y. J., Kim, M. T., Lee, J. E. & Kim, K. B.", "journal": "Journal of Advanced Nursing, 63(5)", "doi": None},
    {"year": 2007, "title": "Evidence-based nursing: effects of a structured nursing program for the health promotion of Korean women with Hwa-Byung", "authors": "Choi, Y. J. & Lee, K. J.", "journal": "Archives of Psychiatric Nursing, 21(1)", "doi": None},
]

COPYRIGHTS = [
    "재난경험자를 위한 마음 치유 기술",
    "재난구호자를 위한 심리적 인명구조술",
    "Psychological Recovery Skills for Disaster Survivors",
    "Psychological Life Support for Disaster Relief Workers",
    "기능성 게임 기반 심리적 응급처치 교육 프로그램 — 지진",
    "기능성 게임 기반 심리적 응급처치 교육 프로그램 — 화재",
    "기능성 게임 기반 심리적 응급처치 교육 프로그램 — 풍수해",
    "기능성 게임 기반 심리적 응급처치 교육 프로그램 — 감염병",
]


def _articles_by_year():
    result = OrderedDict()
    all_pubs = ARTICLES + load_extra_pubs()
    for pub in sorted(all_pubs, key=lambda x: -x["year"]):
        result.setdefault(pub["year"], []).append(pub)
    return result

# ─── Public routes ───────────────────────────────────────────
@app.route("/")
def home():
    recent = sorted(ARTICLES, key=lambda x: -x["year"])[:5]
    news_list = sb("GET","news",params="?published=eq.1&order=created_at.desc") or []
    return render_template("home.html", research_topics=RESEARCH_TOPICS, recent_pubs=recent, news=news_list[:5])

@app.route("/research")
def research():
    extras = {}
    try:
        rows = sb("GET", "research_topics_extra", params="") or []
        for row in (rows if isinstance(rows, list) else []):
            extras[row['key']] = row
    except: pass
    merged = []
    for r in RESEARCH_TOPICS:
        t = dict(r)
        if t['key'] in extras:
            ex = extras[t['key']]
            if ex.get('summary'): t['summary'] = ex['summary']
            if ex.get('detail'):  t['detail']  = ex['detail']
        merged.append(t)
    return render_template("research.html", research_topics=merged, is_researcher='researcher' in session)

@app.route("/team")
def team():
    return render_template("team.html", professor=PROFESSOR, team=TEAM)

@app.route("/publications")
def publications():
    by_year = _articles_by_year()
    return render_template("publications.html", publications_by_year=by_year,
                           pub_years=list(by_year.keys()), copyrights=COPYRIGHTS,
                           current_year=datetime.now().year)

@app.route("/publications/new", methods=["POST"])
@login_required
def publications_new():
    title = request.form.get('title','').strip()
    if not title: return redirect(url_for('publications'))
    try:
        year = int(request.form.get('year', datetime.now().year))
    except ValueError:
        year = datetime.now().year
    doi_raw = request.form.get('doi','').strip()
    entry = {
        'year': year,
        'title': title,
        'authors': request.form.get('authors','').strip(),
        'journal': request.form.get('journal','').strip(),
        'doi': doi_raw if doi_raw else None,
    }
    extra = load_extra_pubs()
    extra.append(entry)
    save_extra_pubs(extra)
    return redirect(url_for('publications'))

@app.route("/apps")
def apps():
    all_apps = APPS + load_extra_apps()
    return render_template("apps.html", apps=all_apps)

@app.route("/apps/new", methods=["POST"])
@login_required
def apps_new():
    name = request.form.get('name','').strip()
    if not name: return redirect(url_for('apps'))
    key = re.sub(r'[^a-z0-9_]','', name.lower().replace(' ','_'))
    if not key: key = 'app'
    extra = load_extra_apps()
    taken = {a['key'] for a in APPS} | {a['key'] for a in extra}
    base, n = key, 2
    while key in taken: key = f"{base}_{n}"; n += 1
    color_map = {
        'teal':   'linear-gradient(135deg,#0F6B6B 0%,#1ec0c0 100%)',
        'blue':   'linear-gradient(135deg,#1a237e 0%,#5c6bc0 100%)',
        'purple': 'linear-gradient(135deg,#4a148c 0%,#9c27b0 100%)',
        'green':  'linear-gradient(135deg,#1b5e20 0%,#66bb6a 100%)',
        'orange': 'linear-gradient(135deg,#e65100 0%,#ff9800 100%)',
    }
    gradient = color_map.get(request.form.get('color','teal'), color_map['teal'])
    app_entry = {'key':key,'icon':request.form.get('icon','📱'),'gradient':gradient,
        'name':name,'full_name':request.form.get('full_name',''),'label':request.form.get('label',''),
        'desc_kr':'','desc_en':request.form.get('desc_en',''),'tags':[],'screenshot':None,
        'content_img_kr':None,'content_img_en':None,'dl_kr':None,'dl_en':None,'subsections':[],
        'created_at':datetime.now().isoformat()[:10]}
    extra.append(app_entry)
    save_extra_apps(extra)
    return redirect(url_for('app_detail', key=key))

@app.route("/apps/<key>")
def app_detail(key):
    all_apps = APPS + load_extra_apps()
    app_data = next((a for a in all_apps if a.get('key')==key), None)
    if not app_data: return redirect(url_for('apps'))
    return render_template("app_detail.html", app=app_data)

@app.route("/contact", methods=["GET","POST"])
def contact():
    success = False
    if request.method == "POST":
        name    = request.form.get("name","").strip()
        email   = request.form.get("email","").strip()
        subject = request.form.get("subject","").strip()
        message = request.form.get("message","").strip()
        if message:
            sb("POST", "contact_messages", data={
                "name": name, "email": email, "subject": subject, "message": message
            })
            success = True
        else:
            flash("메시지를 입력해주세요.")
    return render_template("contact.html", success=success)

@app.route("/gallery")
def gallery():
    events = load_events()
    for ev in events:
        photos = event_photos(ev['key'])
        ev['photo_count'] = len(photos)
        ev['cover_photo'] = ev.get('cover') or (photos[0] if photos else None)
    return render_template("gallery.html", events=events)

@app.route("/gallery/new", methods=["POST"])
@login_required
def gallery_new():
    title = request.form.get('title','').strip()
    if not title: return redirect(url_for('gallery'))
    key = re.sub(r'[^a-z0-9_]','', title.lower().replace(' ','_').replace('(','').replace(')',''))
    if not key: key = 'event'
    events = load_events()
    existing = {e['key'] for e in events}
    base, n = key, 2
    while key in existing: key = f"{base}_{n}"; n += 1
    ev = {'key':key,'title':title,'date':request.form.get('date',''),
          'venue':request.form.get('venue',''),'description':request.form.get('description',''),
          'cover':None,'created_at':datetime.now().isoformat()[:10]}
    events.insert(0, ev)
    save_events(events)
    os.makedirs(os.path.join(GALLERY_FOLDER, key), exist_ok=True)
    return redirect(url_for('gallery_event', key=key))

@app.route("/gallery/<key>")
def gallery_event(key):
    events = load_events()
    ev = next((e for e in events if e['key']==key), None)
    if not ev: return redirect(url_for('gallery'))
    photos = event_photos(key)
    return render_template("gallery_event.html", event=ev, photos=photos)

@app.route("/gallery/<key>/upload", methods=["POST"])
def gallery_event_upload(key):
    events = load_events()
    ev = next((e for e in events if e['key']==key), None)
    if not ev: return redirect(url_for('gallery'))
    folder = os.path.join(GALLERY_FOLDER, key)
    os.makedirs(folder, exist_ok=True)
    changed = False
    for file in request.files.getlist('photo'):
        if file and file.filename and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            if filename:
                file.save(os.path.join(folder, filename))
                if not ev.get('cover'): ev['cover'] = filename; changed = True
    if changed: save_events(events)
    return redirect(url_for('gallery_event', key=key))

@app.route("/gallery/<key>/delete/<filename>", methods=["POST"])
def gallery_event_delete(key, filename):
    safe = secure_filename(filename)
    path = os.path.join(GALLERY_FOLDER, key, safe)
    if os.path.exists(path): os.remove(path)
    events = load_events()
    ev = next((e for e in events if e['key']==key), None)
    if ev and ev.get('cover')==safe:
        remaining = event_photos(key)
        ev['cover'] = remaining[0] if remaining else None
        save_events(events)
    return redirect(url_for('gallery_event', key=key))

@app.route("/gallery/<key>/delete-event", methods=["POST"])
def gallery_delete_event(key):
    events = load_events()
    save_events([e for e in events if e['key']!=key])
    return redirect(url_for('gallery'))

# ─── Auth ────────────────────────────────────────────────────
@app.route("/login", methods=["GET","POST"])
@limiter.limit("10 per minute", methods=["POST"])
def login():
    error = None
    if request.method == "POST":
        email    = request.form.get("email","").strip()
        password = request.form.get("password","").strip()
        stored = _get_account(email)
        if stored and check_password_hash(stored, password):
            session["researcher"] = email
            return redirect(url_for("portal"))
        error = "이메일 또는 비밀번호가 올바르지 않습니다."
    return render_template("login.html", error=error)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("home"))

# 프로젝트 기반 포털
# ════════════════════════════════════════════════════

@app.route("/portal")
@login_required
def portal():
    email = session["researcher"]
    projects = sb("GET","projects",params=f"?researcher_email=eq.{email}&order=created_at.desc") or []
    for p in (projects if isinstance(projects,list) else []):
        pid = p.get("id","")
        pcs = sb("GET","project_participants",params=f"?project_id=eq.{pid}&select=id") or []
        mcs = sb("GET","measurements",params=f"?project_id=eq.{pid}&select=id") or []
        p["participant_count"] = len(pcs) if isinstance(pcs,list) else 0
        p["measurement_count"] = len(mcs) if isinstance(mcs,list) else 0
    all_p = sb("GET","project_participants",params="?select=id") or []
    all_m = sb("GET","measurements",params="?select=id") or []
    return render_template("portal.html",
        researcher=email,
        projects=projects if isinstance(projects,list) else [],
        total_participants=len(all_p) if isinstance(all_p,list) else 0,
        total_measurements=len(all_m) if isinstance(all_m,list) else 0)

@app.route("/portal/projects/new", methods=["POST"])
@login_required
def portal_project_new():
    sb("POST","projects",data={
        "name":            request.form.get("name","").strip(),
        "description":     request.form.get("description","").strip(),
        "researcher_email":session["researcher"],
        "app_type":        request.form.get("app_type","").strip(),
    })
    return redirect(url_for("portal"))

@app.route("/portal/projects/<project_id>")
@login_required
def portal_project(project_id):
    project, err = _require_project_owner(project_id)
    if err: return err
    participants = sb("GET","project_participants",params=f"?project_id=eq.{project_id}&order=enrolled_at.desc") or []
    measurements_raw = sb("GET","measurements",params=f"?project_id=eq.{project_id}&order=measured_at.desc") or []
    variables = sb("GET","project_variables",params=f"?project_id=eq.{project_id}") or []
    p_map = {p["id"]: {"code": p["code"], "group_name": p.get("group_name") or ""} for p in (participants if isinstance(participants,list) else [])}
    for m in (measurements_raw if isinstance(measurements_raw,list) else []):
        info = p_map.get(m.get("participant_id",""), {})
        m["participant_code"] = info.get("code", "")
        m["group_name"] = info.get("group_name", "")
    for p in (participants if isinstance(participants,list) else []):
        p["mcount"] = sum(1 for m in (measurements_raw if isinstance(measurements_raw,list) else []) if m.get("participant_id")==p["id"])
    collaborators = sb("GET", "project_collaborators",
                       params=f"?project_id=eq.{project_id}&order=invited_at.asc") or []
    protocols = sb("GET", "project_protocols",
                   params=f"?project_id=eq.{project_id}&order=sort_order.asc") or []
    is_owner = project.get("researcher_email") == session.get("researcher")
    return render_template("portal_project.html",
        project=project,
        participants=participants if isinstance(participants,list) else [],
        measurements=measurements_raw if isinstance(measurements_raw,list) else [],
        variables=variables if isinstance(variables,list) else [],
        collaborators=collaborators,
        protocols=protocols,
        is_owner=is_owner,
        researcher=session["researcher"])

@app.route("/portal/projects/<project_id>/delete", methods=["POST"])
@login_required
def portal_project_delete(project_id):
    proj, err = _require_project_owner(project_id)
    if err: return err
    pname = proj["name"]
    sb("DELETE","measurements",params=f"?project_id=eq.{project_id}")
    sb("DELETE","project_participants",params=f"?project_id=eq.{project_id}")
    sb("DELETE","project_variables",params=f"?project_id=eq.{project_id}")
    sb("DELETE","projects",params=f"?id=eq.{project_id}")
    audit("delete_project", "projects", project_id, f"name={pname}")
    flash("프로젝트가 삭제됐습니다.")
    return redirect(url_for("portal"))

@app.route("/portal/projects/<project_id>/participants/add", methods=["POST"])
@login_required
def portal_project_add_participant(project_id):
    _, err = _require_project_owner(project_id)
    if err: return err
    code = request.form.get("code","").strip()
    age_raw = request.form.get("age","").strip()
    # Server-side validation
    if not code or len(code) > 50:
        flash("참여자 코드는 1~50자로 입력해주세요."); return redirect(url_for("portal_project", project_id=project_id))
    if not re.match(r'^[A-Za-z0-9_\-가-힣]+$', code):
        flash("코드에 사용할 수 없는 문자가 포함됐어요."); return redirect(url_for("portal_project", project_id=project_id))
    age = None
    if age_raw:
        if not age_raw.isdigit() or not (1 <= int(age_raw) <= 120):
            flash("나이는 1~120 사이 숫자로 입력해주세요."); return redirect(url_for("portal_project", project_id=project_id))
        age = int(age_raw)
    # Duplicate code check
    existing = sb("GET","project_participants",params=f"?project_id=eq.{project_id}&code=eq.{code}")
    if existing:
        flash(f"'{code}' 코드는 이미 이 프로젝트에 존재해요."); return redirect(url_for("portal_project", project_id=project_id))
    sb("POST","project_participants",data={
        "project_id": project_id,
        "code":       code,
        "gender":     request.form.get("gender","").strip() or None,
        "age":        age,
        "group_name": request.form.get("group_name","").strip() or None,
    })
    return redirect(url_for("portal_project", project_id=project_id))

@app.route("/portal/projects/<project_id>/participants/<participant_id>/delete", methods=["POST"])
@login_required
def portal_project_delete_participant(project_id, participant_id):
    parts = sb("GET","project_participants",params=f"?id=eq.{participant_id}")
    code = parts[0]["code"] if parts else participant_id
    sb("DELETE","measurements",params=f"?participant_id=eq.{participant_id}")
    sb("DELETE","project_participants",params=f"?id=eq.{participant_id}")
    audit("delete_participant", "project_participants", participant_id, f"code={code},project={project_id}")
    return redirect(url_for("portal_project", project_id=project_id))

@app.route("/portal/projects/<project_id>/participants/<participant_id>/edit", methods=["POST"])
@login_required
def portal_edit_participant(project_id, participant_id):
    age_raw = request.form.get("age","").strip()
    sb("PATCH","project_participants",
       data={"gender":request.form.get("gender","").strip() or None,
             "age":int(age_raw) if age_raw.isdigit() else None,
             "group_name":request.form.get("group_name","").strip() or None},
       params=f"?id=eq.{participant_id}")
    return redirect(url_for("portal_participant_detail", project_id=project_id, participant_id=participant_id))

@app.route("/portal/projects/<project_id>/participants/<participant_id>")
@login_required
def portal_participant_detail(project_id, participant_id):
    proj, err = _require_project_owner(project_id)
    if err: return err
    p_list = sb("GET","project_participants",params=f"?id=eq.{participant_id}")
    if not p_list or not isinstance(p_list,list): return redirect(url_for("portal_project",project_id=project_id))
    participant = p_list[0]
    measurements = sb("GET","measurements",params=f"?participant_id=eq.{participant_id}&order=measured_at.asc") or []
    variables = sb("GET","project_variables",params=f"?project_id=eq.{project_id}") or []
    # Pre→Post changes (first vs last phase for each variable)
    changes = {}
    # Multi-phase delta transitions: list of {from_phase, to_phase, deltas:{vname:{diff,pct}}}
    phase_deltas = []
    if isinstance(measurements,list) and isinstance(variables,list) and measurements:
        # Simple pre/post (first vs last measurement)
        first_m = measurements[0]
        last_m  = measurements[-1] if len(measurements)>1 else None
        pre_data  = first_m.get("data") or {}
        post_data = last_m.get("data") or {} if last_m else {}
        if last_m and first_m["id"] != last_m["id"]:
            for v in variables:
                vn = v["name"]
                try:
                    pre_val  = float(pre_data.get(vn) or 0)
                    post_val = float(post_data.get(vn) or 0)
                    pct = round((post_val-pre_val)/pre_val*100,1) if pre_val else None
                    changes[vn] = {"pre":pre_val,"post":post_val,
                                   "diff":round(post_val-pre_val,2),"pct":pct,
                                   "label":v.get("label") or vn,"unit":v.get("unit","")}
                except: pass
        # Consecutive phase deltas
        for i in range(len(measurements)-1):
            m_a = measurements[i];  m_b = measurements[i+1]
            d_a = m_a.get("data") or {};  d_b = m_b.get("data") or {}
            entry = {"from_phase": m_a.get("phase") or f"측정{i+1}",
                     "to_phase":   m_b.get("phase") or f"측정{i+2}",
                     "deltas": {}}
            for v in variables:
                vn = v["name"]
                try:
                    va = float(d_a.get(vn) or 0);  vb = float(d_b.get(vn) or 0)
                    pct = round((vb-va)/va*100,1) if va else None
                    entry["deltas"][vn] = {"diff":round(vb-va,2),"pct":pct,
                                           "label":v.get("label") or vn,"unit":v.get("unit","")}
                except: pass
            if entry["deltas"]:
                phase_deltas.append(entry)
    return render_template("portal_participant_detail.html",
        project=proj, participant=participant,
        measurements=measurements if isinstance(measurements,list) else [],
        variables=variables if isinstance(variables,list) else [],
        changes=changes, phase_deltas=phase_deltas,
        researcher=session["researcher"])

@app.route("/portal/projects/<project_id>/measurements/add", methods=["POST"])
@login_required
def portal_project_add_measurement(project_id):
    _, err = _require_project_owner(project_id)
    if err: return err
    variables = sb("GET","project_variables",params=f"?project_id=eq.{project_id}") or []
    if variables and isinstance(variables,list):
        data_dict = {}
        for v in variables:
            val = request.form.get(f"var_{v['name']}","").strip()
            if val:
                try:    data_dict[v["name"]] = float(val) if v.get("var_type")=="number" else val
                except: data_dict[v["name"]] = val
    else:
        raw = request.form.get("data_json","").strip()
        try:    data_dict = json.loads(raw) if raw else {}
        except: data_dict = {"raw": raw}
    sb("POST","measurements",data={
        "project_id":     project_id,
        "participant_id": request.form.get("participant_id","").strip(),
        "phase":          request.form.get("phase","").strip() or None,
        "notes":          request.form.get("notes","").strip() or None,
        "data":           data_dict,
    })
    return redirect(url_for("portal_project", project_id=project_id))

@app.route("/portal/projects/<project_id>/measurements/<measurement_id>/delete", methods=["POST"])
@login_required
def portal_project_delete_measurement(project_id, measurement_id):
    sb("DELETE","measurements",params=f"?id=eq.{measurement_id}")
    audit("delete_measurement", "measurements", measurement_id, f"project={project_id}")
    return redirect(url_for("portal_project", project_id=project_id))

@app.route("/portal/projects/<project_id>/measurements/<measurement_id>/edit", methods=["POST"])
@login_required
def portal_edit_measurement(project_id, measurement_id):
    variables = sb("GET","project_variables",params=f"?project_id=eq.{project_id}") or []
    if variables and isinstance(variables,list):
        data_dict = {}
        for v in variables:
            val = request.form.get(f"var_{v['name']}","").strip()
            if val:
                try:    data_dict[v["name"]] = float(val) if v.get("var_type")=="number" else val
                except: data_dict[v["name"]] = val
    else:
        raw = request.form.get("data_json","").strip()
        try:    data_dict = json.loads(raw) if raw else {}
        except: data_dict = {"raw": raw}
    sb("PATCH","measurements",
       data={"phase":request.form.get("phase","").strip() or None,
             "notes":request.form.get("notes","").strip() or None,
             "data":data_dict},
       params=f"?id=eq.{measurement_id}")
    return redirect(url_for("portal_participant_detail", project_id=project_id,
                            participant_id=request.form.get("participant_id","")))

@app.route("/portal/projects/<project_id>/variables/add", methods=["POST"])
@login_required
def portal_project_add_variable(project_id):
    _, err = _require_project_owner(project_id)
    if err: return err
    sb("POST","project_variables",data={
        "project_id": project_id,
        "name":       request.form.get("name","").strip(),
        "label":      request.form.get("label","").strip() or None,
        "var_type":   request.form.get("var_type","number"),
        "unit":       request.form.get("unit","").strip() or None,
    })
    return redirect(url_for("portal_project", project_id=project_id))

@app.route("/portal/projects/<project_id>/export")
@login_required
def portal_project_export(project_id):
    proj, err = _require_project_owner(project_id)
    if err: return err
    proj_name = proj["name"]
    participants = sb("GET","project_participants",params=f"?project_id=eq.{project_id}") or []
    measurements = sb("GET","measurements",params=f"?project_id=eq.{project_id}&order=measured_at.desc") or []
    p_map = {p["id"]: p for p in (participants if isinstance(participants,list) else [])}
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["participant_code","group","gender","age","phase","measured_at","notes","data"])
    for m in (measurements if isinstance(measurements,list) else []):
        p = p_map.get(m.get("participant_id",""), {})
        writer.writerow([p.get("code",""),p.get("group_name",""),p.get("gender",""),p.get("age",""),
                         m.get("phase",""),m.get("measured_at","")[:10],
                         m.get("notes",""),json.dumps(m.get("data",{}),ensure_ascii=False)])
    output.seek(0)
    safe = proj_name.replace(" ","_")
    return Response(output.getvalue(), mimetype="text/csv",
        headers={"Content-Disposition":f"attachment;filename={safe}_{datetime.now().strftime('%Y%m%d')}.csv"})

@app.route("/portal/projects/<project_id>/export/range")
@login_required
def portal_project_export_range(project_id):
    proj, err = _require_project_owner(project_id)
    if err: return err
    proj_name = proj["name"]
    participants = sb("GET","project_participants",params=f"?project_id=eq.{project_id}") or []
    variables    = sb("GET","project_variables",params=f"?project_id=eq.{project_id}") or []
    measurements = sb("GET","measurements",params=f"?project_id=eq.{project_id}&order=measured_at.asc") or []
    p_map = {p["id"]: p for p in (participants if isinstance(participants,list) else [])}
    # Read filter params
    date_from = request.args.get("date_from","").strip()
    date_to   = request.args.get("date_to","").strip()
    phase_f   = request.args.get("phase","").strip()
    group_f   = request.args.get("group","").strip()
    var_f     = request.args.get("var_name","").strip()
    val_min   = request.args.get("val_min","").strip()
    val_max   = request.args.get("val_max","").strip()
    excl_f    = request.args.get("exclude_excl","1")  # "1" = exclude flagged rows by default
    var_names = [v["name"] for v in (variables if isinstance(variables,list) else [])]
    output = io.StringIO()
    writer = csv.writer(output)
    header = ["participant_code","group","gender","age","phase","measured_at","notes"] + var_names
    writer.writerow(header)
    for m in (measurements if isinstance(measurements,list) else []):
        if excl_f == "1" and m.get("excluded"): continue
        dt = (m.get("measured_at") or "")[:10]
        if date_from and dt < date_from: continue
        if date_to   and dt > date_to:   continue
        if phase_f and m.get("phase","") != phase_f: continue
        p = p_map.get(m.get("participant_id",""), {})
        if group_f and p.get("group_name","") != group_f: continue
        data = m.get("data") or {}
        if var_f and (val_min or val_max):
            try:
                v_val = float(data.get(var_f) or 0)
                if val_min and v_val < float(val_min): continue
                if val_max and v_val > float(val_max): continue
            except: pass
        row = [p.get("code",""), p.get("group_name",""), p.get("gender",""), p.get("age",""),
               m.get("phase",""), dt, m.get("notes","")]
        for vn in var_names:
            row.append(data.get(vn,""))
        writer.writerow(row)
    output.seek(0)
    safe = proj_name.replace(" ","_")
    return Response(output.getvalue(), mimetype="text/csv;charset=utf-8-sig",
        headers={"Content-Disposition":f"attachment;filename={safe}_range_{datetime.now().strftime('%Y%m%d')}.csv"})

@app.route("/portal/projects/<project_id>/upload", methods=["POST"])
@login_required
def portal_project_upload(project_id):
    _, err = _require_project_owner(project_id)
    if err: return err
    f = request.files.get("file")
    if not f or not f.filename:
        flash("파일을 선택해주세요.")
        return redirect(url_for("portal_project", project_id=project_id))
    fname = f.filename.lower()
    rows = []
    tmp_path = None
    try:
        if fname.endswith(".csv"):
            content = f.read().decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(content))
            rows = list(reader)
        elif fname.endswith((".xlsx",".xls")):
            import tempfile, openpyxl
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp_path = tmp.name
                f.save(tmp_path)
            wb = openpyxl.load_workbook(tmp_path, read_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip() for c in next(ws.iter_rows(max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: (str(v).strip() if v is not None else "") for i,v in enumerate(row)})
            wb.close()
        else:
            flash("CSV 또는 Excel 파일만 업로드 가능해요.")
            return redirect(url_for("portal_project", project_id=project_id))
    except Exception as e:
        flash(f"파일 읽기 오류: {e}")
        return redirect(url_for("portal_project", project_id=project_id))
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try: os.unlink(tmp_path)
            except: pass

    variables = sb("GET","project_variables",params=f"?project_id=eq.{project_id}") or []
    var_names = [v["name"] for v in (variables if isinstance(variables,list) else [])]
    var_types = {v["name"]: v.get("var_type","number") for v in (variables if isinstance(variables,list) else [])}
    existing = sb("GET","project_participants",params=f"?project_id=eq.{project_id}") or []
    p_map = {p["code"]: p["id"] for p in (existing if isinstance(existing,list) else [])}
    added_p = added_m = 0
    row_errors = []
    for i, row in enumerate(rows, start=2):
        code = str(row.get("code") or row.get("참여자코드") or "").strip()
        if not code:
            row_errors.append(f"행{i}: 코드 누락"); continue
        if len(code) > 50:
            row_errors.append(f"행{i}: 코드 너무 길어요 ({code[:12]}…)"); continue
        if code not in p_map:
            age_raw = str(row.get("age") or row.get("나이") or "").strip()
            age_val = int(age_raw) if age_raw.isdigit() and 1 <= int(age_raw) <= 120 else None
            new_p = sb("POST","project_participants",data={
                "project_id": project_id, "code": code,
                "gender":     str(row.get("gender") or row.get("성별") or "").strip() or None,
                "age":        age_val,
                "group_name": str(row.get("group") or row.get("그룹") or "").strip() or None,
            })
            if new_p and isinstance(new_p, list):
                p_map[code] = new_p[0]["id"]; added_p += 1
            else:
                row_errors.append(f"행{i}: 참여자 저장 실패 ({code})"); continue
        pid = p_map.get(code)
        if not pid:
            row_errors.append(f"행{i}: 참여자 ID 없음 ({code})"); continue
        phase = str(row.get("phase") or row.get("회차") or "").strip() or None
        notes = str(row.get("notes") or row.get("메모") or "").strip() or None
        data_dict = {}
        skip = {"code","참여자코드","gender","성별","age","나이","group","그룹","group_name","phase","회차","notes","메모"}
        for k, v in row.items():
            if k not in skip and v:
                if var_types.get(k,"number") == "number":
                    try: data_dict[k] = float(v)
                    except: row_errors.append(f"행{i}: {k} 값 '{v}' 숫자 변환 실패")
                else:
                    data_dict[k] = str(v)
        if data_dict:
            sb("POST","measurements",data={"project_id":project_id,"participant_id":pid,
                                           "phase":phase,"notes":notes,"data":data_dict})
            added_m += 1
    msg = f"업로드 완료 — 참여자 {added_p}명, 측정 {added_m}건 추가"
    if row_errors:
        msg += f" / 오류 {len(row_errors)}건: " + "; ".join(row_errors[:5])
        if len(row_errors) > 5: msg += f" 외 {len(row_errors)-5}건"
    flash(msg)
    return redirect(url_for("portal_project", project_id=project_id))

@app.route("/portal/projects/<project_id>/stats")
@login_required
def portal_project_stats(project_id):
    import math
    proj, err = _require_project_owner(project_id)
    if err: return err
    measurements = sb("GET","measurements",params=f"?project_id=eq.{project_id}") or []
    variables    = sb("GET","project_variables",params=f"?project_id=eq.{project_id}") or []
    participants = sb("GET","project_participants",params=f"?project_id=eq.{project_id}") or []
    stats = {}
    if isinstance(measurements,list) and isinstance(variables,list):
        for v in variables:
            vn = v["name"]
            all_vals = []; by_phase = {}
            for m in measurements:
                val = m.get("data",{}).get(vn) if m.get("data") else None
                if val is not None:
                    try:
                        fval=float(val); all_vals.append(fval)
                        ph=m.get("phase","전체") or "전체"
                        by_phase.setdefault(ph,[]).append(fval)
                    except: pass
            def calc(vals):
                if not vals: return {}
                n=len(vals); avg=sum(vals)/n
                sd=math.sqrt(sum((x-avg)**2 for x in vals)/n) if n>1 else 0
                s=sorted(vals); med=s[n//2] if n%2 else (s[n//2-1]+s[n//2])/2
                return {"n":n,"mean":round(avg,2),"sd":round(sd,2),"median":round(med,2),"min":round(min(vals),2),"max":round(max(vals),2)}
            outliers = []
            pid_to_code = {p["id"]: p.get("code","") for p in (participants if isinstance(participants,list) else [])}
            if len(all_vals)>=4:
                s_sorted=sorted(all_vals); q1=s_sorted[len(s_sorted)//4]; q3=s_sorted[3*len(s_sorted)//4]
                iqr=q3-q1; lb,ub=q1-1.5*iqr,q3+1.5*iqr
                for m in measurements:
                    if m.get("excluded"): continue
                    val=m.get("data",{}).get(vn) if m.get("data") else None
                    if val is not None:
                        try:
                            fval=float(val)
                            if fval<lb or fval>ub:
                                outliers.append({
                                    "measurement_id": m["id"],
                                    "participant_code": pid_to_code.get(m.get("participant_id",""),"?"),
                                    "phase": m.get("phase",""),
                                    "value": fval,
                                    "lb": round(lb,2), "ub": round(ub,2)
                                })
                        except: pass
            missing=[]
            phases=list({m.get("phase","") for m in measurements if m.get("phase")})
            for p in (participants if isinstance(participants,list) else []):
                p_ms=[m for m in measurements if m.get("participant_id")==p["id"]]
                for ph in phases:
                    if not any(m.get("phase")==ph and m.get("data",{}).get(vn) is not None for m in p_ms):
                        missing.append({"code":p["code"],"phase":ph})
            # by_group stats (group_name per participant)
            pid_to_group = {p["id"]: (p.get("group_name") or "미지정") for p in (participants if isinstance(participants,list) else [])}
            by_group = {}
            by_group_phase = {}  # {group: {phase: [vals]}}
            for m in measurements:
                val = m.get("data",{}).get(vn) if m.get("data") else None
                if val is not None:
                    try:
                        fval = float(val)
                        grp = pid_to_group.get(m.get("participant_id",""), "미지정")
                        by_group.setdefault(grp, []).append(fval)
                        ph = m.get("phase","전체") or "전체"
                        by_group_phase.setdefault(grp, {}).setdefault(ph, []).append(fval)
                    except: pass
            stats[vn]={"label":v.get("label") or vn,"unit":v.get("unit",""),
                       "overall":calc(all_vals),"by_phase":{ph:calc(vals) for ph,vals in by_phase.items()},
                       "by_group":{grp:calc(vals) for grp,vals in by_group.items()},
                       "by_group_phase":{grp:{ph:calc(vals) for ph,vals in phases.items()} for grp,phases in by_group_phase.items()},
                       "outliers":outliers,"missing":missing}
    # collect all phases and groups for template
    all_phases = sorted({m.get("phase","") for m in (measurements if isinstance(measurements,list) else []) if m.get("phase")})
    all_groups = sorted({(p.get("group_name") or "미지정") for p in (participants if isinstance(participants,list) else [])})
    return render_template("portal_project_stats.html",
        project=proj,stats=stats,variables=variables if isinstance(variables,list) else [],
        all_phases=all_phases,all_groups=all_groups,
        researcher=session["researcher"])

# ── Bulk outlier exclude ──────────────────────────────
@app.route("/portal/projects/<project_id>/outliers/exclude", methods=["POST"])
@login_required
def portal_outliers_exclude(project_id):
    _, err = _require_project_owner(project_id)
    if err: return err
    ids = request.form.getlist("measurement_ids")
    if ids:
        for mid in ids:
            sb("PATCH", "measurements", data={"excluded": 1},
               params=f"?id=eq.{mid}&project_id=eq.{project_id}")
    flash(f"{len(ids)}건의 이상값을 제외 처리했습니다.")
    return redirect(url_for("portal_project_stats", project_id=project_id))

# ── Variable edit / delete ────────────────────────────
@app.route("/portal/projects/<project_id>/variables/<var_id>/edit", methods=["POST"])
@login_required
def portal_variable_edit(project_id, var_id):
    _, err = _require_project_owner(project_id)
    if err: return err
    sb("PATCH", "project_variables",
       data={
           "label":    request.form.get("label","").strip() or None,
           "unit":     request.form.get("unit","").strip() or None,
           "var_type": request.form.get("var_type","number"),
       },
       params=f"?id=eq.{var_id}")
    return redirect(url_for("portal_project", project_id=project_id, _anchor="tab-variables"))

@app.route("/portal/projects/<project_id>/variables/<var_id>/delete", methods=["POST"])
@login_required
def portal_variable_delete(project_id, var_id):
    vrows = sb("GET","project_variables",params=f"?id=eq.{var_id}")
    vname = vrows[0]["name"] if vrows else var_id
    sb("DELETE", "project_variables", params=f"?id=eq.{var_id}")
    audit("delete_variable", "project_variables", var_id, f"name={vname},project={project_id}")
    return redirect(url_for("portal_project", project_id=project_id, _anchor="tab-variables"))

# ── Measurement exclude toggle ────────────────────────
@app.route("/portal/projects/<project_id>/measurements/<m_id>/toggle-exclude", methods=["POST"])
@login_required
def portal_measurement_toggle_exclude(project_id, m_id):
    _, err = _require_project_owner(project_id)
    if err: return err
    rows = sb("GET", "measurements", params=f"?id=eq.{m_id}")
    if rows:
        new_val = 0 if rows[0].get("excluded") else 1
        sb("PATCH", "measurements", data={"excluded": new_val}, params=f"?id=eq.{m_id}")
        audit("toggle_exclude_measurement", "measurements", m_id,
              f"excluded={new_val},project={project_id}")
    return redirect(url_for("portal_project", project_id=project_id, _anchor="tab-measurements"))

# ── Statistical tests page ────────────────────────────
@app.route("/portal/projects/<project_id>/tests")
@login_required
def portal_project_tests(project_id):
    import math
    proj, err = _require_project_owner(project_id)
    if err: return err
    measurements_raw = sb("GET","measurements",params=f"?project_id=eq.{project_id}") or []
    variables    = sb("GET","project_variables",params=f"?project_id=eq.{project_id}") or []
    participants = sb("GET","project_participants",params=f"?project_id=eq.{project_id}") or []

    # Inject group_name into measurements
    pid_to_group = {p["id"]: (p.get("group_name") or "미지정") for p in (participants if isinstance(participants,list) else [])}
    measurements = [m for m in (measurements_raw if isinstance(measurements_raw,list) else []) if not m.get("excluded")]
    for m in measurements:
        m["group_name"] = pid_to_group.get(m.get("participant_id",""), "미지정")

    groups = sorted({m["group_name"] for m in measurements})
    phases = sorted({m.get("phase","") for m in measurements if m.get("phase")})

    results = {}
    if isinstance(variables,list):
        for v in variables:
            vn = v["name"]
            vres = {"label": v.get("label") or vn, "unit": v.get("unit",""), "tests": [], "paired": []}

            # --- Independent tests between every pair of groups ---
            grp_vals = {}
            for m in measurements:
                val = m.get("data",{}).get(vn) if m.get("data") else None
                if val is not None:
                    try: grp_vals.setdefault(m["group_name"],[]).append(float(val))
                    except: pass

            grp_list = list(grp_vals.keys())

            # Normality tests (Shapiro-Wilk per group)
            normality = {}
            for grp, vals in grp_vals.items():
                if 3 <= len(vals) <= 5000:
                    try:
                        sw_stat, sw_p = _scipy_stats.shapiro(vals)
                        normality[grp] = {"stat": round(float(sw_stat),4), "p": round(float(sw_p),4), "normal": sw_p >= 0.05}
                    except: pass
            vres["normality"] = normality

            for i in range(len(grp_list)):
                for j in range(i+1, len(grp_list)):
                    g1, g2 = grp_list[i], grp_list[j]
                    a, b = grp_vals[g1], grp_vals[g2]
                    if len(a) < 2 or len(b) < 2: continue
                    try:
                        t_stat, t_p = _scipy_stats.ttest_ind(a, b, equal_var=False)
                        u_stat, u_p = _scipy_stats.mannwhitneyu(a, b, alternative="two-sided")
                        # Cohen's d
                        pooled_n = len(a)+len(b)
                        pooled_sd = _np.sqrt(((len(a)-1)*_np.var(a,ddof=1) + (len(b)-1)*_np.var(b,ddof=1)) / (pooled_n-2)) if pooled_n > 2 else 0
                        cohens_d = ((_np.mean(a) - _np.mean(b)) / pooled_sd) if pooled_sd > 0 else 0
                        # Bootstrap 95% CI for Cohen's d
                        d_ci = None
                        try:
                            n_boot = 1000
                            boot_d = []
                            rng = _np.random.default_rng(42)
                            for _ in range(n_boot):
                                ba = rng.choice(a, size=len(a), replace=True)
                                bb = rng.choice(b, size=len(b), replace=True)
                                ps = float(_np.sqrt(((len(ba)-1)*_np.var(ba,ddof=1) + (len(bb)-1)*_np.var(bb,ddof=1)) / (len(ba)+len(bb)-2)))
                                if ps > 0: boot_d.append((float(_np.mean(ba))-float(_np.mean(bb)))/ps)
                            if boot_d:
                                d_ci = [round(float(_np.percentile(boot_d,2.5)),3), round(float(_np.percentile(boot_d,97.5)),3)]
                        except: pass
                        # Rank-biserial r for Mann-Whitney
                        r_rb = 1 - 2*u_stat / (len(a)*len(b))
                        vres["tests"].append({
                            "g1": g1, "g2": g2,
                            "n1": len(a), "n2": len(b),
                            "mean1": round(float(_np.mean(a)),3), "mean2": round(float(_np.mean(b)),3),
                            "sd1": round(float(_np.std(a,ddof=1)),3), "sd2": round(float(_np.std(b,ddof=1)),3),
                            "t": round(float(t_stat),3), "t_p": round(float(t_p),4),
                            "U": round(float(u_stat),1), "u_p": round(float(u_p),4),
                            "d": round(float(cohens_d),3), "r": round(float(r_rb),3),
                            "t_sig": t_p < 0.05, "u_sig": u_p < 0.05,
                            "d_ci": d_ci,
                        })
                    except Exception as e:
                        print(f"test error {vn} {g1} vs {g2}: {e}")

            # ANOVA / Kruskal-Wallis for 3+ groups
            if len(grp_list) >= 3:
                valid_groups = [g for g in grp_list if len(grp_vals[g]) >= 2]
                all_groups_data = [grp_vals[g] for g in valid_groups]
                if len(all_groups_data) >= 3:
                    try:
                        f_stat, f_p = _scipy_stats.f_oneway(*all_groups_data)
                        h_stat, h_p = _scipy_stats.kruskal(*all_groups_data)
                        grand = [x for g in all_groups_data for x in g]
                        grand_mean = float(_np.mean(grand))
                        ss_between = sum(len(g)*(float(_np.mean(g))-grand_mean)**2 for g in all_groups_data)
                        ss_total = sum((x-grand_mean)**2 for x in grand)
                        eta2 = round(ss_between/ss_total, 3) if ss_total > 0 else 0
                        # Tukey HSD post-hoc (only if ANOVA is significant)
                        tukey_pairs = []
                        if f_p < 0.05:
                            try:
                                tukey_res = _scipy_stats.tukey_hsd(*all_groups_data)
                                for ti in range(len(valid_groups)):
                                    for tj in range(ti+1, len(valid_groups)):
                                        gi, gj = valid_groups[ti], valid_groups[tj]
                                        tval = tukey_res.statistic[ti][tj]
                                        tpval = tukey_res.pvalue[ti][tj]
                                        tukey_pairs.append({
                                            "g1": gi, "g2": gj,
                                            "stat": round(float(tval),3),
                                            "p": round(float(tpval),4),
                                            "sig": float(tpval) < 0.05,
                                        })
                            except Exception as te:
                                print(f"tukey error {vn}: {te}")
                        vres["anova"] = {
                            "groups": valid_groups,
                            "F": round(float(f_stat),3), "f_p": round(float(f_p),4), "f_sig": f_p < 0.05,
                            "H": round(float(h_stat),3), "h_p": round(float(h_p),4), "h_sig": h_p < 0.05,
                            "eta2": eta2,
                            "tukey": tukey_pairs,
                        }
                    except Exception as e:
                        print(f"anova error {vn}: {e}")

            # --- Paired t-test (pre vs post within same participant) ---
            if len(phases) >= 2:
                ph_pairs = [(phases[i], phases[i+1]) for i in range(len(phases)-1)]
                for pre_ph, post_ph in ph_pairs:
                    paired_pre, paired_post = [], []
                    for p in (participants if isinstance(participants,list) else []):
                        pre_m  = [m for m in measurements if m.get("participant_id")==p["id"] and m.get("phase")==pre_ph]
                        post_m = [m for m in measurements if m.get("participant_id")==p["id"] and m.get("phase")==post_ph]
                        if pre_m and post_m:
                            pv = pre_m[0].get("data",{}).get(vn)
                            ov = post_m[0].get("data",{}).get(vn)
                            if pv is not None and ov is not None:
                                try: paired_pre.append(float(pv)); paired_post.append(float(ov))
                                except: pass
                    if len(paired_pre) >= 2:
                        try:
                            t_s, t_p = _scipy_stats.ttest_rel(paired_pre, paired_post)
                            diffs = [b-a for a,b in zip(paired_pre,paired_post)]
                            sd_d = float(_np.std(diffs,ddof=1))
                            d_paired = float(_np.mean(diffs)) / sd_d if sd_d > 0 else 0
                            vres["paired"].append({
                                "pre_phase": pre_ph, "post_phase": post_ph,
                                "n": len(paired_pre),
                                "mean_pre": round(float(_np.mean(paired_pre)),3),
                                "mean_post": round(float(_np.mean(paired_post)),3),
                                "mean_diff": round(float(_np.mean(diffs)),3),
                                "sd_diff": round(sd_d,3),
                                "t": round(float(t_s),3), "p": round(float(t_p),4),
                                "d": round(d_paired,3),
                                "sig": t_p < 0.05,
                            })
                        except Exception as e:
                            print(f"paired test error {vn}: {e}")
            results[vn] = vres

    # --- Correlation matrix ---
    corr_matrix = None
    corr_labels = []
    if isinstance(variables,list) and len(variables) >= 2:
        num_vars = [v for v in variables if v.get("var_type","number") == "number"]
        if len(num_vars) >= 2:
            corr_labels = [v.get("label") or v["name"] for v in num_vars]
            col_data = []
            for v in num_vars:
                col_data.append([float(m.get("data",{}).get(v["name"])) for m in measurements
                                  if m.get("data",{}).get(v["name"]) is not None
                                  and str(m.get("data",{}).get(v["name"],"")).replace(".","",1).lstrip("-").isdigit()])
            # Use pairwise complete obs approach
            n_vars = len(num_vars)
            matrix = [[None]*n_vars for _ in range(n_vars)]
            for i in range(n_vars):
                for j in range(n_vars):
                    if i == j:
                        matrix[i][j] = {"r": 1.0, "p": 0.0, "sig": False}
                    else:
                        # Get paired values (rows where both have data)
                        pairs = []
                        for m in measurements:
                            if not m.get("data"): continue
                            vi = m["data"].get(num_vars[i]["name"])
                            vj = m["data"].get(num_vars[j]["name"])
                            if vi is not None and vj is not None:
                                try: pairs.append((float(vi), float(vj)))
                                except: pass
                        if len(pairs) >= 3:
                            xs = [p[0] for p in pairs]
                            ys = [p[1] for p in pairs]
                            try:
                                r, p_val = _scipy_stats.pearsonr(xs, ys)
                                matrix[i][j] = {"r": round(float(r), 3), "p": round(float(p_val), 4), "sig": p_val < 0.05}
                            except: matrix[i][j] = None
            corr_matrix = matrix

    return render_template("portal_project_tests.html",
        project=proj,
        results=results,
        variables=variables if isinstance(variables,list) else [],
        corr_matrix=corr_matrix,
        corr_labels=corr_labels,
        groups=groups,
        phases=phases,
        researcher=session["researcher"])

@app.route("/portal/projects/<project_id>/tests/export")
@login_required
def portal_project_tests_export(project_id):
    """Export statistical test results as CSV."""
    _, err = _require_project_owner(project_id)
    if err: return err
    # Re-run the test logic (lightweight version returning only flat rows)
    measurements_raw = sb("GET","measurements",params=f"?project_id=eq.{project_id}") or []
    variables    = sb("GET","project_variables",params=f"?project_id=eq.{project_id}") or []
    participants = sb("GET","project_participants",params=f"?project_id=eq.{project_id}") or []
    proj = sb("GET","projects",params=f"?id=eq.{project_id}")
    proj_name = proj[0]["name"] if proj else "project"
    pid_to_group = {p["id"]: (p.get("group_name") or "미지정") for p in (participants if isinstance(participants,list) else [])}
    measurements = [m for m in (measurements_raw if isinstance(measurements_raw,list) else []) if not m.get("excluded")]
    for m in measurements:
        m["group_name"] = pid_to_group.get(m.get("participant_id",""), "미지정")
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["변수","비교유형","그룹1","그룹2","N1","N2","평균1","평균2","t/U/F","p값","유의","Cohen's d","효과크기"])
    for v in (variables if isinstance(variables,list) else []):
        vn = v["name"]; label = v.get("label") or vn
        grp_vals = {}
        for m in measurements:
            val = m.get("data",{}).get(vn)
            if val is not None:
                try: grp_vals.setdefault(m["group_name"],[]).append(float(val))
                except: pass
        grp_list = list(grp_vals.keys())
        for i in range(len(grp_list)):
            for j in range(i+1, len(grp_list)):
                g1, g2 = grp_list[i], grp_list[j]
                a, b = grp_vals[g1], grp_vals[g2]
                if len(a) < 2 or len(b) < 2: continue
                try:
                    t_s, t_p = _scipy_stats.ttest_ind(a, b, equal_var=False)
                    n = len(a)+len(b)
                    ps = _np.sqrt(((len(a)-1)*_np.var(a,ddof=1) + (len(b)-1)*_np.var(b,ddof=1)) / (n-2)) if n > 2 else 0
                    d = (float(_np.mean(a))-float(_np.mean(b)))/float(ps) if ps > 0 else 0
                    eff = "large" if abs(d)>=0.8 else "medium" if abs(d)>=0.5 else "small"
                    writer.writerow([label,"독립표본t",g1,g2,len(a),len(b),round(float(_np.mean(a)),3),round(float(_np.mean(b)),3),round(float(t_s),3),round(float(t_p),4),"*" if t_p<0.05 else "",round(float(d),3),eff])
                except: pass
        if len(grp_list) >= 3:
            gdata = [grp_vals[g] for g in grp_list if len(grp_vals[g]) >= 2]
            if len(gdata) >= 3:
                try:
                    f_s, f_p = _scipy_stats.f_oneway(*gdata)
                    writer.writerow([label,"ANOVA","(전체)","",sum(len(g) for g in gdata),"",
                                     "","",round(float(f_s),3),round(float(f_p),4),"*" if f_p<0.05 else "","",""])
                except: pass
    output.seek(0)
    safe = proj_name.replace(" ","_")
    return Response(output.getvalue(), mimetype="text/csv;charset=utf-8-sig",
        headers={"Content-Disposition":f"attachment;filename={safe}_stats_{datetime.now().strftime('%Y%m%d')}.csv"})

@app.route("/portal/files")
@login_required
def portal_files():
    email = session["researcher"]
    files = sb("GET", "portal_files", params=f"?researcher_email=eq.{email}&order=uploaded_at.desc") or []
    projects = sb("GET", "projects", params=f"?researcher_email=eq.{email}&order=created_at.desc") or []
    # Build project name lookup
    proj_map = {p["id"]: p["name"] for p in (projects if isinstance(projects, list) else [])}
    for f in (files if isinstance(files, list) else []):
        f["project_name"] = proj_map.get(f.get("project_id"), "")
    return render_template("portal_files.html", files=files, projects=projects if isinstance(projects, list) else [], researcher=email)

@app.route("/portal/files/upload", methods=["POST"])
@login_required
def portal_files_upload():
    f = request.files.get("file")
    if not f or not f.filename:
        flash("파일을 선택해주세요.")
        return redirect(url_for("portal_files"))
    original = f.filename
    safe = secure_filename(original)
    if not safe:
        flash("유효하지 않은 파일명입니다.")
        return redirect(url_for("portal_files"))
    uid = str(_uuid.uuid4())
    ext = os.path.splitext(safe)[1]
    stored = uid + ext
    os.makedirs(PORTAL_FILES_DIR, exist_ok=True)
    f.save(os.path.join(PORTAL_FILES_DIR, stored))
    size = os.path.getsize(os.path.join(PORTAL_FILES_DIR, stored))
    sb("POST", "portal_files", data={
        "filename": stored, "original_name": original,
        "size": size, "researcher_email": session["researcher"],
        "project_id": request.form.get("project_id") or None,
    })
    flash(f"'{original}' 업로드 완료")
    return redirect(url_for("portal_files"))

@app.route("/portal/files/<file_id>/download")
@login_required
def portal_files_download(file_id):
    rows = sb("GET", "portal_files", params=f"?id=eq.{file_id}")
    if not rows: return redirect(url_for("portal_files"))
    row = rows[0]
    if row.get("researcher_email") != session["researcher"]:
        flash("접근 권한이 없습니다."); return redirect(url_for("portal_files"))
    path = os.path.join(PORTAL_FILES_DIR, row["filename"])
    if not os.path.exists(path): return redirect(url_for("portal_files"))
    from flask import send_file
    return send_file(path, as_attachment=True, download_name=row["original_name"])

@app.route("/portal/files/<file_id>/delete", methods=["POST"])
@login_required
def portal_files_delete(file_id):
    rows = sb("GET", "portal_files", params=f"?id=eq.{file_id}")
    if rows and rows[0].get("researcher_email") == session["researcher"]:
        path = os.path.join(PORTAL_FILES_DIR, rows[0]["filename"])
        if os.path.exists(path): os.remove(path)
        sb("DELETE", "portal_files", params=f"?id=eq.{file_id}")
    return redirect(url_for("portal_files"))

@app.route("/portal/settings")
@login_required
def portal_settings():
    return render_template("portal_settings.html", researcher=session["researcher"])

@app.route("/portal/settings/password", methods=["POST"])
@login_required
def portal_change_password():
    current=request.form.get("current","").strip()
    new_pw=request.form.get("new_pw","").strip()
    confirm=request.form.get("confirm","").strip()
    email=session["researcher"]
    stored = _get_account(email)
    if not stored or not check_password_hash(stored, current): flash("현재 비밀번호가 틀렸어요."); return redirect(url_for("portal_settings"))
    if new_pw!=confirm: flash("새 비밀번호가 일치하지 않아요."); return redirect(url_for("portal_settings"))
    if len(new_pw)<6: flash("비밀번호는 6자 이상이어야 해요."); return redirect(url_for("portal_settings"))
    _set_account(email, new_pw)
    flash("비밀번호가 변경됐어요.")
    return redirect(url_for("portal_settings"))

@app.route("/portal/settings/add-account", methods=["POST"])
@login_required
def portal_add_account():
    new_email=request.form.get("email","").strip()
    new_pw=request.form.get("password","").strip()
    if not new_email or not new_pw: flash("이메일과 비밀번호를 입력해주세요."); return redirect(url_for("portal_settings"))
    if _account_exists(new_email): flash("이미 존재하는 계정이에요."); return redirect(url_for("portal_settings"))
    _set_account(new_email, new_pw)
    flash(f"{new_email} 계정이 추가됐어요.")
    return redirect(url_for("portal_settings"))

@app.route("/portal/settings/backup")
@login_required
def portal_db_backup():
    import shutil
    backup_buf = io.BytesIO()
    # Use SQLite's backup API for a consistent snapshot
    src = sqlite3.connect(DB_PATH)
    dst = sqlite3.connect(":memory:")
    src.backup(dst)
    src.close()
    # Write in-memory DB to buffer
    for chunk in dst.iterdump():
        pass  # iterdump is text; use serialize instead (Python 3.11+) or file copy
    dst.close()
    # Simpler: just send the file directly (safe because WAL is enabled)
    with open(DB_PATH, "rb") as f:
        data = f.read()
    fname = f"portal_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
    return Response(data, mimetype="application/octet-stream",
        headers={"Content-Disposition": f"attachment;filename={fname}"})

@app.route("/portal/settings/restore", methods=["POST"])
@login_required
def portal_db_restore():
    f = request.files.get("backup_file")
    if not f or not f.filename.endswith(".db"):
        flash("올바른 .db 파일을 선택해주세요.")
        return redirect(url_for("portal_settings"))
    # Validate it's a SQLite file
    header = f.read(16)
    if not header.startswith(b"SQLite format 3"):
        flash("유효한 SQLite 데이터베이스 파일이 아닙니다.")
        return redirect(url_for("portal_settings"))
    f.seek(0)
    # Write to a temp file then validate before replacing
    import tempfile, shutil
    with tempfile.NamedTemporaryFile(delete=False, suffix=".db") as tmp:
        f.save(tmp.name)
        try:
            test_conn = sqlite3.connect(tmp.name)
            test_conn.execute("SELECT name FROM sqlite_master LIMIT 1").fetchone()
            test_conn.close()
        except Exception as e:
            os.unlink(tmp.name)
            flash(f"복원 실패: DB 파일 검증 오류 — {e}")
            return redirect(url_for("portal_settings"))
        shutil.copy2(tmp.name, DB_PATH)
        os.unlink(tmp.name)
    flash("데이터베이스가 성공적으로 복원됐어요. 페이지를 새로고침해주세요.")
    return redirect(url_for("portal_settings"))

@app.route("/portal/merge")
@login_required
def portal_merge():
    email=session["researcher"]
    projects=sb("GET","projects",params=f"?researcher_email=eq.{email}&order=created_at.desc") or []
    return render_template("portal_merge.html",researcher=email,
        projects=projects if isinstance(projects,list) else [])

@app.route("/portal/merge/preview", methods=["POST"])
@login_required
def portal_merge_preview():
    p1_id=request.form.get("project1_id",""); p2_id=request.form.get("project2_id","")
    if not p1_id or not p2_id or p1_id==p2_id: flash("서로 다른 프로젝트를 선택해주세요."); return redirect(url_for("portal_merge"))
    p1=sb("GET","projects",params=f"?id=eq.{p1_id}"); p2=sb("GET","projects",params=f"?id=eq.{p2_id}")
    p1_parts=sb("GET","project_participants",params=f"?project_id=eq.{p1_id}") or []
    p2_parts=sb("GET","project_participants",params=f"?project_id=eq.{p2_id}") or []
    p1_codes={p["code"] for p in (p1_parts if isinstance(p1_parts,list) else [])}
    p2_codes={p["code"] for p in (p2_parts if isinstance(p2_parts,list) else [])}
    return render_template("portal_merge.html",researcher=session["researcher"],
        projects=sb("GET","projects",params=f"?researcher_email=eq.{session['researcher']}&order=created_at.desc") or [],
        preview=True,project1=p1[0] if p1 else {},project2=p2[0] if p2 else {},
        common=sorted(p1_codes&p2_codes),only_p1=sorted(p1_codes-p2_codes),only_p2=sorted(p2_codes-p1_codes),
        p1_id=p1_id,p2_id=p2_id)

@app.route("/portal/merge/execute", methods=["POST"])
@login_required
def portal_merge_execute():
    p1_id=request.form.get("p1_id",""); p2_id=request.form.get("p2_id","")
    new_name=request.form.get("new_name","").strip() or "병합 프로젝트"
    new_proj=sb("POST","projects",data={"name":new_name,"description":f"병합: {p1_id[:8]}+{p2_id[:8]}",
        "researcher_email":session["researcher"]})
    if not new_proj or not isinstance(new_proj,list): flash("병합 실패"); return redirect(url_for("portal_merge"))
    new_pid=new_proj[0]["id"]
    all_vars={}
    for src in [p1_id,p2_id]:
        for v in (sb("GET","project_variables",params=f"?project_id=eq.{src}") or []):
            all_vars[v["name"]]=v
    for vn,v in all_vars.items():
        sb("POST","project_variables",data={"project_id":new_pid,"name":vn,
            "label":v.get("label"),"var_type":v.get("var_type","number"),"unit":v.get("unit")})
    p_code_map={}
    for src in [p1_id,p2_id]:
        for p in (sb("GET","project_participants",params=f"?project_id=eq.{src}") or []):
            if p["code"] not in p_code_map:
                new_p=sb("POST","project_participants",data={"project_id":new_pid,"code":p["code"],
                    "gender":p.get("gender"),"age":p.get("age"),"group_name":p.get("group_name")})
                if new_p and isinstance(new_p,list): p_code_map[p["code"]]=new_p[0]["id"]
    for src_pid in [p1_id,p2_id]:
        src_parts=sb("GET","project_participants",params=f"?project_id=eq.{src_pid}") or []
        src_map={p["id"]:p["code"] for p in (src_parts if isinstance(src_parts,list) else [])}
        tag="P1" if src_pid==p1_id else "P2"
        for m in (sb("GET","measurements",params=f"?project_id=eq.{src_pid}") or []):
            code=src_map.get(m.get("participant_id",""),"")
            new_pid2=p_code_map.get(code)
            if not new_pid2: continue
            sb("POST","measurements",data={"project_id":new_pid,"participant_id":new_pid2,
                "phase":f"[{tag}] {m.get('phase','')}" if m.get("phase") else tag,
                "notes":m.get("notes"),"data":m.get("data",{})})
    flash(f"병합 완료! '{new_name}' 프로젝트가 생성됐어요.")
    return redirect(url_for("portal_project", project_id=new_pid))

@app.route("/portal/all/participants")
@login_required
def portal_all_participants():
    participants=sb("GET","project_participants",params="?select=*,projects(name)&order=enrolled_at.desc") or []
    return render_template("portal_participants.html",researcher=session["researcher"],
        participants=participants if isinstance(participants,list) else [])

@app.route("/portal/all/sessions")
@login_required
def portal_all_sessions():
    session_list = sb("GET", "sessions", params="?order=received_at.desc") or []
    if session_list:
        all_parts = sb("GET", "project_participants", params="?select=id,code") or []
        p_map = {p["id"]: p["code"] for p in (all_parts if isinstance(all_parts, list) else [])}
        for s in session_list:
            s["participant_code"] = p_map.get(s.get("participant_id", ""), "")
    return render_template("portal_sessions.html", researcher=session["researcher"],
        sessions=session_list)

@app.route("/api/sessions", methods=["POST"])
def api_receive_session():
    if request.headers.get("X-API-Key","")!=os.getenv("APP_API_KEY","tsl-app-key-2025"):
        return jsonify({"error":"Unauthorized"}),401
    payload = request.get_json(silent=True) or {}
    sid = str(_uuid.uuid4())
    sb("POST", "sessions", data={
        "id":             sid,
        "project_id":     payload.get("project_id"),
        "participant_id": payload.get("participant_id"),
        "phase":          payload.get("phase", ""),
        "notes":          payload.get("notes", ""),
        "data":           payload.get("data", {}),
    })
    return jsonify({"ok": True, "id": sid}), 201

# ── Project meta edit ────────────────────────────────
@app.route("/portal/projects/<project_id>/edit", methods=["POST"])
@login_required
def portal_project_edit(project_id):
    _, err = _require_project_owner(project_id)
    if err: return err
    new_name = request.form.get("name","").strip()
    sb("PATCH", "projects",
       data={"name": new_name,
             "description": request.form.get("description","").strip(),
             "app_type": request.form.get("app_type","").strip()},
       params=f"?id=eq.{project_id}")
    audit("edit_project", "projects", project_id, f"name={new_name}")
    flash("프로젝트 정보가 수정됐어요.")
    return redirect(url_for("portal_project", project_id=project_id))

# ── Project clone ─────────────────────────────────────
@app.route("/portal/projects/<project_id>/clone", methods=["POST"])
@login_required
def portal_project_clone(project_id):
    proj, err = _require_project_owner(project_id)
    if err: return err
    variables = sb("GET","project_variables",params=f"?project_id=eq.{project_id}") or []
    new_id = str(_uuid.uuid4())
    sb("POST","projects",data={
        "id": new_id,
        "name": proj["name"] + " (복사)",
        "description": proj.get("description",""),
        "app_type": proj.get("app_type",""),
        "researcher_email": session["researcher"],
    })
    for v in (variables if isinstance(variables,list) else []):
        sb("POST","project_variables",data={
            "project_id": new_id,
            "name": v["name"], "label": v.get("label"), "var_type": v.get("var_type","number"), "unit": v.get("unit")
        })
    flash(f"'{proj['name']}' 프로젝트가 복사됐어요. (변수 설정만 복사, 데이터 제외)")
    return redirect(url_for("portal_project", project_id=new_id))

# ── Excel export ──────────────────────────────────────
@app.route("/portal/projects/<project_id>/export/excel")
@login_required
def portal_project_export_excel(project_id):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    proj, err = _require_project_owner(project_id)
    if err: return err
    participants = sb("GET","project_participants",params=f"?project_id=eq.{project_id}") or []
    measurements = sb("GET","measurements",params=f"?project_id=eq.{project_id}") or []
    variables    = sb("GET","project_variables",params=f"?project_id=eq.{project_id}") or []
    p_map = {p["id"]: p for p in (participants if isinstance(participants,list) else [])}
    var_names = [v["name"] for v in (variables if isinstance(variables,list) else [])]
    var_labels = {v["name"]: (v.get("label") or v["name"]) for v in (variables if isinstance(variables,list) else [])}

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0F6B6B")

    def style_header(ws, headers):
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = max(12, len(h)+2)

    # Sheet 1: All data
    ws1 = wb.active; ws1.title = "전체 데이터"
    hdrs = ["참여자코드","그룹","성별","나이","회차","측정일","메모"] + [var_labels.get(n,n) for n in var_names]
    style_header(ws1, hdrs)
    for m in (measurements if isinstance(measurements,list) else []):
        if m.get("excluded"): continue
        p = p_map.get(m.get("participant_id",""), {})
        row = [p.get("code",""), p.get("group_name",""), p.get("gender",""), p.get("age",""),
               m.get("phase",""), m.get("measured_at","")[:10], m.get("notes","")]
        for n in var_names:
            val = m.get("data",{}).get(n) if m.get("data") else None
            row.append(val)
        ws1.append(row)

    # Sheet 2: Group x Phase means
    ws2 = wb.create_sheet("그룹×회차 평균")
    phases = sorted({m.get("phase","") for m in (measurements if isinstance(measurements,list) else []) if m.get("phase") and not m.get("excluded")})
    groups = sorted({p.get("group_name","미지정") for p in (participants if isinstance(participants,list) else [])})
    s2_hdrs = ["변수","그룹"] + phases
    style_header(ws2, s2_hdrs)
    for vn in var_names:
        for grp in groups:
            row = [var_labels.get(vn,vn), grp]
            for ph in phases:
                vals = []
                for m in (measurements if isinstance(measurements,list) else []):
                    if m.get("excluded"): continue
                    p = p_map.get(m.get("participant_id",""),{})
                    if (p.get("group_name") or "미지정") == grp and m.get("phase") == ph:
                        v = m.get("data",{}).get(vn) if m.get("data") else None
                        if v is not None:
                            try: vals.append(float(v))
                            except: pass
                row.append(round(sum(vals)/len(vals),3) if vals else "")
            ws2.append(row)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    proj_name = proj["name"].replace("/","_").replace("\\","_")
    return Response(buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={proj_name}.xlsx"})

# ── Measurement edit ──────────────────────────────────
@app.route("/portal/projects/<project_id>/measurements/<m_id>/edit", methods=["POST"])
@login_required
def portal_measurement_edit(project_id, m_id):
    _, err = _require_project_owner(project_id)
    if err: return err
    rows = sb("GET","measurements",params=f"?id=eq.{m_id}")
    if not rows: return redirect(url_for("portal_project", project_id=project_id))
    variables = sb("GET","project_variables",params=f"?project_id=eq.{project_id}") or []
    data_dict = {}
    if isinstance(variables, list):
        for v in variables:
            val = request.form.get(f"var_{v['name']}", "").strip()
            if val:
                try: data_dict[v['name']] = float(val)
                except: data_dict[v['name']] = val
    sb("PATCH","measurements",
       data={"phase": request.form.get("phase","").strip() or None,
             "notes": request.form.get("notes","").strip() or None,
             "data": data_dict},
       params=f"?id=eq.{m_id}")
    return redirect(url_for("portal_project", project_id=project_id, _anchor="tab-measurements"))

# ── Participant bulk group change ─────────────────────
@app.route("/portal/projects/<project_id>/participants/bulk-group", methods=["POST"])
@login_required
def portal_participants_bulk_group(project_id):
    ids = request.form.getlist("participant_ids")
    group_name = request.form.get("group_name","").strip() or None
    for pid in ids:
        sb("PATCH","project_participants",data={"group_name": group_name},params=f"?id=eq.{pid}")
    flash(f"{len(ids)}명의 그룹이 변경됐어요.")
    return redirect(url_for("portal_project", project_id=project_id, _anchor="tab-participants"))

# ── News management ───────────────────────────────────
@app.route("/portal/news")
@login_required
def portal_news():
    email = session["researcher"]
    news_list = sb("GET","news",params="?order=created_at.desc") or []
    return render_template("portal_news.html", news=news_list, researcher=email)

@app.route("/portal/news/new", methods=["POST"])
@login_required
def portal_news_new():
    sb("POST","news",data={
        "title": request.form.get("title","").strip(),
        "content": request.form.get("content","").strip(),
        "researcher_email": session["researcher"],
        "published": 1 if request.form.get("published") else 0,
    })
    flash("공지가 등록됐어요.")
    return redirect(url_for("portal_news"))

@app.route("/portal/news/<news_id>/edit", methods=["POST"])
@login_required
def portal_news_edit(news_id):
    sb("PATCH","news",
       data={"title": request.form.get("title","").strip(),
             "content": request.form.get("content","").strip(),
             "published": 1 if request.form.get("published") else 0},
       params=f"?id=eq.{news_id}")
    flash("공지가 수정됐어요.")
    return redirect(url_for("portal_news"))

@app.route("/portal/news/<news_id>/delete", methods=["POST"])
@login_required
def portal_news_delete(news_id):
    sb("DELETE","news",params=f"?id=eq.{news_id}")
    flash("공지가 삭제됐어요.")
    return redirect(url_for("portal_news"))

# ── API documentation ─────────────────────────────────
@app.route("/api/docs")
@login_required
def api_docs():
    api_key = os.getenv("APP_API_KEY","tsl-app-key-2025")
    return render_template("api_docs.html", researcher=session["researcher"], api_key=api_key)

# ── Research topic inline edit ────────────────────────
@app.route("/portal/research/<key>/edit", methods=["POST"])
@login_required
def portal_research_edit(key):
    summary = request.form.get("summary","").strip()
    detail  = request.form.get("detail","").strip()
    sb("POST", "research_topics_extra",
       data={"key": key, "summary": summary, "detail": detail},
       upsert=True)
    flash("연구 소개가 수정됐어요.")
    return redirect(url_for("research") + f"#{key}")

# ── Contact messages inbox ────────────────────────────
@app.route("/portal/contacts")
@login_required
def portal_contacts():
    msgs = sb("GET", "contact_messages", params="?order=created_at.desc") or []
    return render_template("portal_contacts.html", messages=msgs, researcher=session["researcher"])

@app.route("/portal/audit")
@login_required
def portal_audit():
    PAGE_SIZE = 50
    page = max(1, request.args.get("page", 1, type=int))
    total = _sb_count("audit_log")
    offset = (page - 1) * PAGE_SIZE
    logs = sb("GET", "audit_log",
              params=f"?order=created_at.desc&limit={PAGE_SIZE}&offset={offset}") or []
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    return render_template("portal_audit.html", logs=logs, researcher=session["researcher"],
                           page=page, total_pages=total_pages, total=total)

# ── Collaboration ──────────────────────────────────────
@app.route("/portal/projects/<project_id>/collaborators/add", methods=["POST"])
@login_required
def portal_project_add_collaborator(project_id):
    proj, err = _require_project_owner(project_id)
    if err: return err
    # Only the owner can invite collaborators
    if proj.get("researcher_email") != session.get("researcher"):
        flash("프로젝트 소유자만 협력자를 초대할 수 있어요.")
        return redirect(url_for("portal_project", project_id=project_id))
    email = (request.form.get("collaborator_email") or "").strip().lower()
    role  = request.form.get("role", "viewer")
    if not email:
        flash("초대할 연구자 이메일을 입력해 주세요.")
        return redirect(url_for("portal_project", project_id=project_id))
    if email == session.get("researcher"):
        flash("본인을 협력자로 초대할 수 없어요.")
        return redirect(url_for("portal_project", project_id=project_id))
    existing = sb("GET", "project_collaborators",
                  params=f"?project_id=eq.{project_id}&researcher_email=eq.{email}&select=id")
    if existing:
        flash(f"{email}는 이미 협력자로 등록되어 있어요.")
        return redirect(url_for("portal_project", project_id=project_id))
    sb("POST", "project_collaborators", data={
        "project_id": project_id, "researcher_email": email, "role": role
    })
    audit("collaborator_add", "project_collaborators", project_id,
          after={"email": email, "role": role})
    flash(f"{email}를 {role} 역할로 초대했어요.")
    return redirect(url_for("portal_project", project_id=project_id))

@app.route("/portal/projects/<project_id>/collaborators/remove", methods=["POST"])
@login_required
def portal_project_remove_collaborator(project_id):
    proj, err = _require_project_owner(project_id)
    if err: return err
    if proj.get("researcher_email") != session.get("researcher"):
        flash("프로젝트 소유자만 협력자를 제거할 수 있어요.")
        return redirect(url_for("portal_project", project_id=project_id))
    email = (request.form.get("collaborator_email") or "").strip().lower()
    sb("DELETE", "project_collaborators",
       params=f"?project_id=eq.{project_id}&researcher_email=eq.{email}")
    audit("collaborator_remove", "project_collaborators", project_id, before={"email": email})
    flash(f"{email}를 협력자 목록에서 제거했어요.")
    return redirect(url_for("portal_project", project_id=project_id))

# ── Protocol phases ────────────────────────────────────
@app.route("/portal/projects/<project_id>/protocols/save", methods=["POST"])
@login_required
def portal_project_protocols_save(project_id):
    proj, err = _require_project_owner(project_id)
    if err: return err
    phases_raw = request.form.get("phases", "")
    phases = [p.strip() for p in phases_raw.split(",") if p.strip()]
    sb("DELETE", "project_protocols", params=f"?project_id=eq.{project_id}")
    for i, phase in enumerate(phases):
        sb("POST", "project_protocols", data={
            "project_id": project_id, "phase_name": phase, "sort_order": i
        })
    audit("protocol_save", "project_protocols", project_id, after={"phases": phases})
    flash("프로토콜 단계가 저장되었어요.")
    return redirect(url_for("portal_project", project_id=project_id) + "#tab-settings")

# ── Password reset ─────────────────────────────────────
@app.route("/reset-password", methods=["GET","POST"])
def reset_password_request():
    if request.method == "POST":
        email = (request.form.get("email") or "").strip().lower()
        acct = sb("GET", "accounts", params=f"?email=eq.{email}&select=email")
        if acct:
            token = str(_uuid.uuid4())
            expires = (datetime.now() + timedelta(hours=2)).strftime("%Y-%m-%dT%H:%M:%S")
            sb("DELETE", "reset_tokens", params=f"?email=eq.{email}")
            sb("POST", "reset_tokens", data={
                "token": token, "email": email, "expires_at": expires
            })
            reset_url = url_for("reset_password_confirm", token=token, _external=True)
            flash(f"비밀번호 재설정 링크가 생성되었어요. (개발 환경: {reset_url})")
        else:
            flash("해당 이메일로 등록된 계정이 없어요.")
        return redirect(url_for("reset_password_request"))
    return render_template("reset_password_request.html")

@app.route("/reset-password/<token>", methods=["GET","POST"])
def reset_password_confirm(token):
    rows = sb("GET", "reset_tokens", params=f"?token=eq.{token}&select=email,expires_at")
    if not rows:
        flash("유효하지 않거나 이미 사용된 링크예요.")
        return redirect(url_for("login"))
    email, expires_at = rows[0]["email"], rows[0]["expires_at"]
    if datetime.now() > datetime.strptime(expires_at, "%Y-%m-%dT%H:%M:%S"):
        flash("링크가 만료되었어요. 다시 요청해 주세요.")
        return redirect(url_for("reset_password_request"))
    if request.method == "POST":
        pw = request.form.get("password", "")
        if len(pw) < 8:
            flash("비밀번호는 8자 이상이어야 해요.")
            return render_template("reset_password_confirm.html", token=token)
        sb("PATCH", "accounts",
           data={"password": generate_password_hash(pw, method='pbkdf2:sha256')},
           params=f"?email=eq.{email}")
        sb("DELETE", "reset_tokens", params=f"?token=eq.{token}")
        flash("비밀번호가 재설정되었어요. 로그인해 주세요.")
        return redirect(url_for("login"))
    return render_template("reset_password_confirm.html", token=token)

if __name__ == "__main__":
    app.run(debug=False)
